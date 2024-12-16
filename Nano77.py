import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from openpyxl.styles import Border, Side
import math
import string
import numpy as np 
import tkinter as tk
from tkinter import ttk
from PIL import Image
import customtkinter as ctk
from ttkthemes import ThemedTk
from tkinter import messagebox
import logging.config
from concurrent.futures import thread
from selenium.webdriver.common.keys import Keys
from seleniumwire import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from selenium_stealth import stealth
from icecream import ic
# import argparse
import requests
from bs4 import BeautifulSoup as bs
import datetime
import logging
import random
import sys
import threading
import time
# import traceback
import shutil, getpass, requests, zipfile, os, re#, wget
import re
import copy

STAT_LOGGING_INTERVAL = 60  # Seconds
THREAD_SHUTDOWN_TIMEOUT = 3  # Seconds
thread_start_delay = 0


def generate_log_filename():
    now = datetime.datetime.now()
    now_str = now.strftime('%Y_%m_%d_%H_%M_%S')
    log_dir = os.path.join(os.path.normpath(os.getcwd() + os.sep), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    global log_fname
    log_fname = os.path.join(log_dir, now_str) + '.txt'
    return log_fname
logging.basicConfig(
    level=logging.DEBUG,  # Log WARNING and above.
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(generate_log_filename()),
    ]
    
)

class MyFilter(logging.Filter):
    def __init__(self, param=None):
        super(MyFilter, self).__init__()
        self.param = param

    def filter(self, record):
        if self.param is None:
            allow = True
        else:
            allow = self.param not in record.msg
        if allow:
            record.msg = 'changed: ' + record.msg
        return allow
    
LOGGING = {
    'version': 1,
    'filters': {
        'myfilter': {
            '()': MyFilter,
            'param': 'request',
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.FileHandler',
            'filters': ['myfilter'],
        }
    },
    'root': {
        'level': 'DEBUG',
        'handlers': ['console']
    },
}

class SimulationThread(threading.Thread):
    def __init__(self, simulation, thread_id, url, proxy, combo, run_time):
        super(SimulationThread, self).__init__(name=f'SimulationThread-{thread_id}')
        self.simulation = simulation
        self.thread_id = thread_id
        self.url = url
        self.proxy = proxy
        self.combo = combo
        self.requested_stop = False
        self.failed = False
        self.run_time = run_time

    def run(self):
        logging.getLogger("seleniumwire").setLevel(logging.WARNING)
        logging.info(f'Starting thread {self.thread_id}')
        try:
            self.automate(self.thread_id, self.url, self.proxy, self.combo, self.run_time)

        except Exception as e:
            self.failed = True
            error_message = str(e).lower()
            if error_message.find('connection was forcibly closed') != -1:
                # 10054 handler
                logging.info(f'Detected error 10054 in thread {self.thread_id}. Restarting the thread...')
                self.simulation.restart_thread(old_thread=self)
            elif error_message.find('err_tunnel_connection_failed') != -1:
                logging.info(f'Proxy {self.proxy} isn\'t working properly, \n shutting down thread')
                open('non_working_proxy.txt', 'a').write(f'{self.proxy}')
                thread._shutdown
            if error_message.find('no such window'):
                logging.info('Couldn\'t find window')
                thread._shutdown
            else:
                # Generic error handler.
                # This call also logs a stack trace.
                logging.exception(f'Thread {self.thread_id} exited with an error.')
    def stop(self):
        self.requested_stop = True

    def automate(self, id, url, proxy, emails, run_time):
        #show_loading()
        # try:
        #     username, password = combo.split(":")
        # except:
        #     username = combo
        #     password = "3334444"
        lock.acquire()
        wireoptions = {
        # 'proxy': {
        #     'http':'http://%s' %proxy,
        #     'https': 'https://%s' %proxy,
        #     'no_proxy': 'localhost,127.0.0.1'
        #         }
        }
        options = Options()
        options.add_argument("start-maximized")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--mute-audio")
        options.add_argument("--disable-blink-features=AutomationControlled")
        # options.add_extension('active.crx')
        # options.add_extension('css.crx')
        # scriptDirectory = pathlib.Path().absolute()
        # directory = ''.join(random.choice(string.ascii_letters + string.digits) for i in range(8))
        # directory = "chrome_"+ directory
        # options.add_argument(f'--user-data-dir={scriptDirectory}\{directory}')
        s = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=s, options=options, seleniumwire_options=wireoptions)
        stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
        )
        max_goals = 11
        leagues, proxy = proxy["leagues"], proxy["proxies_file"]

        try:
            gameids = []
            lock.release()
            website = "https:/www.goaloo18.com" + emails
            m_url = 'https://www.goaloo18.com/football'
            print(f"\033[92m{website}\033[0m")
            driver.get(website)
            driver.add_cookie({"name": "Time_Zone", "value": "10"})
            driver.add_cookie({"name": "FilterOptionFix", "value": "0"})
            if leagues:
                driver.add_cookie({"name": "orderby", "value": "league"})
            else:
                driver.add_cookie({"name": "orderby", "value": "time"})
            driver.refresh()
            time.sleep(3)
            games = driver.find_elements(By.CSS_SELECTOR, "tr[leaindex][index][id^='tr1']")
            act_data = []
            data = {'Time':'', 'Date': '', 'League':'', 'Home vs Guest':'','Goal Value A':'','Goal Value H':'','Goal Cost Home':'','Goal Cost Away':'', 'Score(Finished)':'', 'Over':'', 'STAKE POOL': '', 'Early':'', 'SUM':'','Live':''}
            for team_score_index in range(1, max_goals):
                data[f'goal{team_score_index}'] = ''
            print(f"\033[93m{len(games)}\033[0m")
            # Click on an element to open a new tab
            original_window_handle = driver.current_window_handle
            for game in games:
                game_time = game.find_element(By.CLASS_NAME, "time").text
                game_time = datetime.datetime.strptime(game_time, "%H:%M").time()
                if url['start_time'] <= game_time <= url['end_time']:
                    gameids.append(game.get_attribute('id').split('_')[1])
                else:
                    continue
            #print(gameids)
            if proxy == 0:
                proxy = len(gameids)
            print(f'trying to calculate data for {proxy} games')
            working_count = 0
            count = 0
            cv_home = []
            GC_home = []
            cv_away = []
            GC_away = []
            t1_avg = []
            t2_avg = []
            last_matches1 = []
            last_matches2 = []
            bad_games = []
            home_goals = []
            cc = 0
            try:
                for gameid in gameids:
                    try:
                        if count >= proxy:
                            df = pd.DataFrame(data)
                            df.to_excel(f'{date.strip()}.xlsx', index=False)
                            #print(f'I hit my {proxy} game count limit, data has been stored to an excel file in the current directory')
                            driver.quit()
                            input('press enter to exit')
                            sys.exit()
                        count+=1
                        try:
                            print(f'analyzing game number {gameid}')
                            driver.get(f'{m_url}/match/live-{gameid}')
                            time.sleep(2)
                            league = driver.find_element(By.XPATH, '//span[@class="sclassLink"]').text
                        except NoSuchElementException:
                            try:    
                                league = driver.find_element(By.CSS_SELECTOR, '.nosclassLink').text
                            except:
                                driver.refresh()
                                league = driver.find_element(By.XPATH, "//*[@id='fbheader']/div[1]/span[1]/span").text
                        except:
                            continue

                        home_team = driver.find_elements(By.XPATH, '//div[@class="sclassName"]')[0].text
                        guest_team = driver.find_elements(By.XPATH, '//div[@class="sclassName"]')[1].text
                        date, match_time, week_day = driver.find_element(By.XPATH, '//span[@class="time"]').text.split(' ')
                        goals = []
                        score = []
                        
                        print("-----")
                        if 'Finished' in driver.find_element(By.XPATH, "//*[@id='mScore']").text:
                            goal_A = driver.find_element(By.XPATH, '//*[@id="mScore"]/div/div[1]').text
                            goal_B = driver.find_element(By.XPATH, '//*[@id="mScore"]/div/div[3]').text
                            score.append(f"{goal_A} x {goal_B}")
                            rows = driver.find_elements(By.XPATH, '//table[@class="team-table-other ky"]/tbody/tr')
                            for row in rows:
                                if row.find_elements(By.XPATH, './/img[@alt="Goal"]') or row.find_elements(By.XPATH, './/img[@alt="Penalty scored"]') or row.find_elements(By.XPATH, './/img[@alt="Own goal"]'):
                                    goal_time = row.find_element(By.XPATH, './/td/b').text
                                    goals.append(goal_time)
                        else:
                            goal_A = "NA"
                            goal_B = "NA"    
                        goals.reverse()
                        driver.get(f'{m_url}/match/over-under-odds-{gameid}')
                        time.sleep(3)
                        companys = driver.find_elements(By.CSS_SELECTOR,"tr.tb-bgcolor")
                        companys += driver.find_elements(By.CSS_SELECTOR,"tr.tb-bgcolor1")
                        time.sleep(1)
                        try:
                            for com in companys:
                                if "Interwetten" == com.find_element(By.CSS_SELECTOR,"td.rb").text:
                                    time.sleep(1)
                                    first_odds = com.find_elements(By.TAG_NAME,"td")[1].text
                                    odds_goals = com.find_elements(By.TAG_NAME,"td")[2].text
                        except:
                            first_odds = 0
                            odds_goals = 0
                            pass
						
                        bet365_early_data_deep_copied = "0.0 / 0.0 / 0.0"
                        bet365_live_data_deep_copied = "0.0 / 0.0 / 0.0"

                        try:
                            driver.get(f'{m_url}/{league}/live-{gameid}')
                            time.sleep(4)
                            early_live_str = driver.find_element(By.XPATH, '//*[@id="oddsDiv_8"]/table/tbody/tr[4]').text
                            new_values = early_live_str.split()
                            bet365_early_data_deep_copied = " / ".join(new_values[1:4])
                            bet365_live_data_deep_copied = " / ".join(new_values[4:7])
                            time.sleep(1)

                        except:
                            try:
                                driver.refresh()
                                time.sleep(3)
                                early_live_str = driver.find_element(By.XPATH, '//*[@id="oddsDiv_8"]/table/tbody/tr[4]').text
                                new_values = early_live_str.split()
                                bet365_early_data_deep_copied = " / ".join(new_values[1:4])
                                bet365_live_data_deep_copied = " / ".join(new_values[4:7])
                                time.sleep(1)
                            except:
                                pass
                            
                        working_count+=1

                        date_deep_copied = str(date)
                        match_time_deep_copied = match_time
                        league_deep_copied = str(league)
                        home_team_deep_copied = str(f'{home_team} VS {guest_team}')
                        goal_A_deep_copied = str(f"{goal_A} x {goal_B}")
                        first_odds_deep_copied = str(first_odds)
                        odds_goals_deep_copied = str(odds_goals)
                        numbers = bet365_early_data_deep_copied.split('/')
                        
                        if (float(numbers[0]) + float(numbers[1]) + float(numbers[2])) == 0:
                            STAKE_POOL_1 = 0
                            STAKE_POOL_2 = 0
                        else:
                            STAKE_POOL_1 = (float(numbers[0])/float(numbers[1])) / 2
                            STAKE_POOL_2 = (float(numbers[2])/float(numbers[1])) + STAKE_POOL_1
                        
                        # Convert each substring to float and sum them
                        total_sum = sum(float(num) for num in numbers)
                        result_ = total_sum / 3
                        rounded_result = round(result_, 5)

                        try:   
                            response = requests.get(f'{m_url}/match/h2h-{gameid}')
                            from time import sleep
                            sleep(5)
                        except:
                            print('cant open URL')
                        
                        soup = BeautifulSoup(response.content, "html.parser")
                        porletP6_element = soup.find(id="porletP6")
                        
                        x  = porletP6_element.find(id="table_v1")

                        y  = porletP6_element.find(id="table_v2")
                        rows_in_table_v2 = x.find_all("tr")
                        rows_in_table_v1 = y.find_all("tr")
                        good_data = True
                        if len(rows_in_table_v1) !=25 or len(rows_in_table_v2) != 25:
                            good_data = False
                            bad_games.append(gameid)

                        print('analyse_success') 
                        #print(date_deep_copied,match_time_deep_copied,league_deep_copied,home_team_deep_copied,goal_A_deep_copied,first_odds_deep_copied,odds_goals_deep_copied,bet365_early_data_deep_copied,bet365_live_data_deep_copied)
                        home = home_team_deep_copied.split("VS")[0]
                        away = home_team_deep_copied.split("VS")[1]
                        GH = goal_A_deep_copied.split("x")[0]
                        GA = goal_A_deep_copied.split("x")[1]
                        if  good_data == True  and first_odds_deep_copied != "-" :
                            cc += 1
                            print('Now on GAME : ',cc,home,away)
                            data['Time'] = match_time_deep_copied
                            data['Date'] = date_deep_copied
                            data['League'] =league_deep_copied
                            data['Home'] = home
                            data['Away'] = away
                            data['GH'] = GH
                            data['GA'] = GA
                            data['Over'] = first_odds_deep_copied
                            data['STAKE POOL'] = ( STAKE_POOL_2 + rounded_result) * 1.75
                            data['Early'] = bet365_early_data_deep_copied
                            data['SUM'] = rounded_result
                            data['Live'] = bet365_live_data_deep_copied
                         
                            values = data['Early'].split('/')
                            value1 = float(values[0].strip())
                            value2 = float(values[1].strip())
                            value3 = float(values[2].strip())
                            if (value1 + value2 + value3) == 0:
                                data['Prob.home'] = 0
                                data['Prob.Away'] = 0
                            else:
                                data['Prob.home'] = str(1/value1)
                                data['Prob.Away'] = str(1/value3)

                            for team_score_index in range(1, max_goals):
                                try:
                                    gol = copy.deepcopy(goals[team_score_index-1])
                                    data[f'goal{team_score_index}'] = gol
                                except IndexError:
                                    data[f'goal{team_score_index}'] = 'NA'
                            #if count%10==0:
                                #print(f'calculated {count} out of {len(gameids)}, {working_count} had Bet365 data {count-working_count} Didn\'t have required data and where skipped')
                            
                            act_data.append(data.copy())
                            driver.get(f'{m_url}/match/h2h-{gameid}#porletP6')
                            
                            porletP5 = driver.find_element(By.ID,"porletP6")
                            table_v1 = porletP5.find_element(By.ID,"table_v1")
                            table_v1_ = {'Team':'','Score':'','W/L':'','AVGH':'','AVGA':'','HGD':'','AGD':'','SD Home':'','SD Away':'','CV Home':'','CV Away':'',"AGH":''}
                            total_goals_t1 = 0
                            
                            goals_table = []
                            sqr_goals_table = []
                            valid_t1_matchs = []
                            SD = 0
                            home_goals = []
                            button = driver.find_element(By.ID, "cb_sos1")
                            button.click()
                            z =10
                            for y in range(1,z):
                                td = table_v1.find_element(By.ID,"tr1_"+str(y))
                                if len(td.find_elements(By.TAG_NAME,"td")[3].text) == 0:
                                    z=z+1
                                else:
                                    home_goals.append(td.find_elements(By.TAG_NAME,"td")[3].text)
                                    print('valeur: ',td.find_elements(By.TAG_NAME,"td")[3].text,y )
                            home_goals = home_goals[0:5]     
                            adx = 0
                            for elmt in home_goals:
                                adx += int(elmt[0:1])
                            
                            #print('adx',adx/5)
                            button.click()
                            time.sleep(1)
                            AGH = 0
                            for x in range(1,6):
                                try:
                                    try:
                                        td = table_v1.find_element(By.ID,"tr1_"+str(x))
                                        if table_v1.find_element(By.TAG_NAME,"a").text  == td.find_elements(By.TAG_NAME,"td")[2].text:
                                            table_v1_['Team'] ='(H) ' + table_v1.find_element(By.TAG_NAME,"a").text
                                            AGH += int(td.find_elements(By.TAG_NAME,"td")[3].text[0])
                                        else:
                                            table_v1_['Team'] = table_v1.find_element(By.TAG_NAME,"a").text + ' (A)'
                                            AGH += int(td.find_elements(By.TAG_NAME,"td")[3].text[2:3])
                                        table_v1_['Score'] =  td.find_elements(By.TAG_NAME,"td")[3].text
                                        table_v1_['W/L'] = td.find_elements(By.TAG_NAME,"td")[9].text
                                    except:
                                        table_v1_['Team'] = "No"
                                        table_v1_['Score'] = '1-1(0-0)'
                                        table_v1_['W/L'] = 'D'

									#Updated by Wassim Karaouli
                                    try:
                                        table_v1_score_value0 = int(table_v1_['Score'][0])
                                        table_v1_score_value2 = int(table_v1_['Score'][2])
                                    except:
                                        table_v1_score_value0 = 0
                                        table_v1_score_value2 = 0

                                    if x == 1 or x == 2:
                                        last_matches1.append([table_v1_score_value0, table_v1_score_value2])

                                    total_goals_t1 += table_v1_score_value0 + table_v1_score_value2
                                    goals_table.append(table_v1_score_value0 + table_v1_score_value2)
                                    
                                    if x == 5 and len(goals_table) == 5:
                                        table_v1_['AGH'] = str(AGH/5) 
                                        t1_avg.append(total_goals_t1 /5)
                                        xx = (total_goals_t1 / 5) * float(data['Prob.Away'])
                                        table_v1_['Goal Cost Home'] = str(float(data['Prob.home']) / (total_goals_t1 / 5))
                                        table_v1_['Goal Value H'] = str(xx)
                                    if x == 5 and len(goals_table) == 5:
                                        table_v1_["AVGH"] = str(total_goals_t1/5)
                                        
                                        op = 0.0
                                        for team_score_index in range(1,5):
                                            op = abs((goals_table[team_score_index])-(total_goals_t1/5))
                                            op = math.sqrt(op)
                                            op = op / 5
                                            sqr_goals_table.append(op)
                                        
                                        sd = sum(sqr_goals_table) /(total_goals_t1/5)
                                        GC_home.append(str(float(data['Prob.home']) / (total_goals_t1 / 5)))
                                        cv_home.append(str(sd/(total_goals_t1/5)))
                                        table_v1_["SD Home"] = str(sd)
                                        table_v1_["CV Home"] = str(sd/(total_goals_t1/5))
                                    
                                    act_data.append(table_v1_.copy())
                                except Exception as e:
                                    pass
                            
                            table_v2 = porletP5.find_element(By.ID,"table_v2")
                            table_v2_ = {'Team':'','Score':'','W/L':'','AVGH':'','AVGA':'','HGD':'','AGD':'','SD Home':'','SD Away':'','CV Home':'','CV Away':'','AGA':''}
                            total_goals_t2 = 0
                            goals_table2 = []
                            SD = 0
                            sqr_goals_table2 = []
                            home_goals = []
                            button = driver.find_element(By.ID, "cb_sos2")
                            button.click()
                            z =10
                            for y in range(1,z):
                                td = table_v2.find_element(By.ID,"tr2_"+str(y))
                                if len(td.find_elements(By.TAG_NAME,"td")[3].text) == 0:
                                    z=z+1
                                else:
                                    home_goals.append(td.find_elements(By.TAG_NAME,"td")[3].text)

                            home_goals = home_goals[0:5]     
                            adx = 0
                            for elmt in home_goals:
                                adx += int(elmt[2:3])

                            button.click()
                            time.sleep(2)
                            AGA = 0
                            for x in range(1,6):
                                try:
                                    try:
                                        td = table_v2.find_element(By.ID,"tr2_"+str(x))
                                        if table_v2.find_element(By.TAG_NAME,"a").text  == td.find_elements(By.TAG_NAME,"td")[2].text:
                                            table_v2_['Team'] ='(H) ' + table_v2.find_element(By.TAG_NAME,"a").text
                                            AGA += int(td.find_elements(By.TAG_NAME,"td")[3].text[0])
                                            
                                        else:
                                            AGA += int(td.find_elements(By.TAG_NAME,"td")[3].text[2:3])
                                            table_v2_['Team'] = table_v2.find_element(By.TAG_NAME,"a").text  + ' (A)'
                                        
                                        table_v2_['Score'] =  td.find_elements(By.TAG_NAME,"td")[3].text
                                        table_v2_['W/L'] = td.find_elements(By.TAG_NAME,"td")[9].text
                                    except:
                                         table_v2_['Team'] = 'No'
                                         table_v2_['Score'] ='1-1(0-0)'
                                         table_v2_['W/L'] = 'D'

                                    try:
                                         table_v2_score_value0 = int(table_v2_['Score'][0])
                                         table_v2_score_value2 = int(table_v2_['Score'][2])
                                    except:
                                         table_v2_score_value0 = 0
                                         table_v2_score_value2 = 0

                                    if x == 1 or x == 2:
                                        last_matches2.append([table_v2_score_value0, table_v2_score_value2])

                                    total_goals_t2 += table_v2_score_value0 + table_v2_score_value2
                                    goals_table2.append(table_v2_score_value0 + table_v2_score_value2)
                                    if x == 5 and len(goals_table2) == 5:
                                        
                                        table_v2_['Goal Cost Away'] = str(float(data['Prob.Away']) / (total_goals_t2 / 5))
                                        xx = (total_goals_t2 / 5) * float(data['Prob.home'])
                                        t2_avg.append(total_goals_t2 / 5)
                                        table_v2_['Goal Value A'] = str(xx)                                    

                                    if x == 5 and len(goals_table2) == 5 :
                                        table_v2_['AGA'] = str(AGA/5) 
                                        table_v2_["AVGA"] = str(total_goals_t2/5)
                                        
                                        op = 0.0
                                        for team_score_index in range(1,5):
                                            op = abs((goals_table2[team_score_index])-(total_goals_t2/5))
                                            op = math.sqrt(op)
                                            op = op / 5
                                            sqr_goals_table2.append(op)
                                                                                    
                                        sd = sum(sqr_goals_table2) /(total_goals_t2/5)
                                        GC_away.append(str(float(data['Prob.Away']) / (total_goals_t2 / 5)))
                                        cv_away.append(str(sd/(total_goals_t2/5)))
                                        table_v2_["SD Away"] = str(sd)
                                        table_v2_["CV Away"] = str(sd/(total_goals_t2/5))                             
                                    act_data.append(table_v2_.copy())
                                except Exception as e:
                                    pass
                            
                            last_mathces = []
                            for team_score_index in range(0,len(last_matches1),2):
                                last_mathces.append([last_matches1[team_score_index],last_matches1[team_score_index+1]])     
                    except Exception as e:
                        pass
            except Exception as e:
                pass
            try:
                df = pd.DataFrame(act_data)
            
                # Updated by Wassim Karaouli (Arrange columns and Rows)
                new_sheet_order = ['Time', 'Date', 'League', 'Home','Away','GH','GA','AVGH', 'AVGA','AGH','AGA','HGD','AGD','Goal Value A', 'Goal Value H', 'Goal Cost Home', 'Goal Cost Away', 'CV Home','CV Away','Prob.home', 'Prob.Away','SD Home','SD Away','Early', 'SUM', 'Live','goal1', 'goal2', 'goal3', 'goal4', 'goal5', 'goal6', 'goal7', 'goal8', 'goal9', 'goal10', 'Over', 'STAKE POOL', 'Team', 'Score', 'W/L']
                df = df.reindex(columns=new_sheet_order)
                df['CV Home'] = df['CV Home'].shift(-5)
                df['Goal Value A'] = df['Goal Value A'].shift(-10)
                df['Goal Cost Away'] = df['Goal Cost Away'].shift(-10)
                df['Goal Cost Home'] = df['Goal Cost Home'].shift(-5)
                df['Goal Value H'] = df['Goal Value H'].shift(-5)
                df['CV Away'] = df['CV Away'].shift(-10)
                df['SD Home'] = df['SD Home'].shift(-5)
                df['SD Away'] = df['SD Away'].shift(-10)
                df['AGA'] = df['AGA'].shift(-10)
                df['AGH'] = df['AGH'].shift(-5)
                df['AVGA'] = df['AVGA'].shift(-10)
                df['AVGH'] = df['AVGH'].shift(-5)
                df['AGD'] = df['AGD'].shift(-10)
                df['HGD'] = df['HGD'].shift(-5)
                # Goal Cost Home, Goal Cost Away, Prob. Home, Prob. Away, SD Home, SD Away, CV Home & CV Away
                GCH = []
                GCA = []
                PH = []
                PA = []
                SDH = []
                GVA = []
                GVH = []
                AVGH = []
                AVGA = []
                AGA = []
                AGH = []
                SDA  = []
                CVH = []
                CVA = []
                HGD = []
                AGD = []
                EARLY = []

                PROV_AH = False
                PROV_SK = False
                PROV_NX = False
                PROV_DBT = False
                PROV_BK = False
                PROV_BK09 = False
                PROV_D3RD = False
                PROV_D5TH = False
                PROV_LTN1ST = False
                PROV_LN7 = False
                PROV_3NR = False
                PROV_HDN = False
                PROV_3DD = False
                PROV_F09 = False
                PROV_PTN = False
                PROV_HCO = False
                PROV_QM = False
                PROV_SCA = False
                PROV_SCR = False
                PROV_TTC = False
                PROV_ECC = False
                PROV_PTN = False
                PROV_TWIS = False
                PROV_GDSZ = False
                PROV_HQO = False
                PROV_THU = False
                PROV_CEO = False
                PROV_COO = False
                PROV_HSS = False
                PROV_VLN = False
                PROV_HOC = False
                PROV_BNG = False
                PROV_DP = False
                PROV_IMC = False
                PROV_TIG = False
                PROV_EAG = False
                PROV_CYC = False
                PROV_NK = False
                PROV_FTF = False
                PROV_NFD = False
                PROV_OWO = False
                PROV_BFM = False
                PROV_NN = False
                PROV_NED = False
                PROV_ASY = False
                PROV_CIS = False
                PROV_AOO = False
                PROV_CIN = False
                PROV_HOY = False
                PROV_WHP = False
                PROV_PAV = False
                PROV_D1R = False
                PROV_D77 = False

                for team_score_index in range(len(df['Goal Value H'])):
                    if isinstance(df['Goal Value H'][team_score_index],str) and df['Goal Value H'][team_score_index]!="" :
                        GVH.append(df['Goal Value H'][team_score_index])
                for team_score_index in range(len(df['Goal Value A'])):
                    if isinstance(df['Goal Value A'][team_score_index],str) and df['Goal Value A'][team_score_index]!="" :
                        GVA.append(df['Goal Value A'][team_score_index])        
                for team_score_index in range(len(df['Goal Cost Home'])):
                    if isinstance(df['Goal Cost Home'][team_score_index],str) and df['Goal Cost Home'][team_score_index]!="" :
                        GCH.append(df['Goal Cost Home'][team_score_index])
                for team_score_index in range(len(df['Goal Cost Away'])):
                    if isinstance(df['Goal Cost Away'][team_score_index],str) and df['Goal Cost Away'][team_score_index]!="" :
                        GCA.append(df['Goal Cost Away'][team_score_index])
                for team_score_index in range(len(df['Prob.home'])):
                    if isinstance(df['Prob.home'][team_score_index],str) and df['Prob.home'][team_score_index]!="" :
                        PH.append(df['Prob.home'][team_score_index])
                for team_score_index in range(len(df['Prob.Away'])):
                    if isinstance(df['Prob.Away'][team_score_index],str) and df['Prob.Away'][team_score_index]!="" :
                        PA.append(df['Prob.Away'][team_score_index])
                for team_score_index in range(len(df['SD Home'])):
                    if isinstance(df['SD Home'][team_score_index],str) and df['SD Home'][team_score_index]!="" :
                        SDH.append(df['SD Home'][team_score_index])
                for team_score_index in range(len(df['SD Away'])):
                    if isinstance(df['SD Away'][team_score_index],str) and df['SD Away'][team_score_index]!="" :
                        SDA.append(df['SD Away'][team_score_index])
                for team_score_index in range(len(df['CV Home'])):
                    if isinstance(df['CV Home'][team_score_index],str) and df['CV Home'][team_score_index]!="" :
                        CVH.append(df['CV Home'][team_score_index])
                for team_score_index in range(len(df['CV Away'])):
                    if isinstance(df['CV Away'][team_score_index],str) and df['CV Away'][team_score_index]!="" :
                        CVA.append(df['CV Away'][team_score_index])
                for team_score_index in range(len(df['AVGH'])):
                    if isinstance(df['AVGH'][team_score_index],str) and df['AVGH'][team_score_index]!="" :
                        AVGH.append(df['AVGH'][team_score_index])
                for team_score_index in range(len(df['AVGA'])):
                    if isinstance(df['AVGA'][team_score_index],str) and df['AVGA'][team_score_index]!="" :
                        AVGA.append(df['AVGA'][team_score_index])
                for team_score_index in range(len(df['AGA'])):
                    if isinstance(df['AGA'][team_score_index],str) and df['AGA'][team_score_index]!="" :
                        AGA.append(df['AGA'][team_score_index])        
                for team_score_index in range(len(df['AGH'])):
                    if isinstance(df['AGH'][team_score_index],str) and df['AGH'][team_score_index]!="" :
                        AGH.append(df['AGH'][team_score_index]) 
                
                for team_score_index in range(len(df['Early'])):
                    if isinstance(df['Early'][team_score_index],str) and df['Early'][team_score_index]!="" :
                        EARLY.append(df['Early'][team_score_index])
                t =[GVA,GVH, GCH,GCA,CVH ,CVA
                ,PH
                ,PA
                ,SDH
                ,SDA
                
                ]
    
                # Check if there is data
                if not df.empty:
                    # Create a new workbook
                    wb = Workbook()
                    ws = wb.active

                    # Append headers separately to ensure they are included
                    header_row = df.columns.tolist()
                    ws.append(header_row)
                    def find(b):
                        for i in range(len(b)):
                            b[i] = b[i][2:4]
                        val = b[0]
                        x = b.count(val)
                        for i in range(1,len(b)):
                            if x < b.count(b[i]):
                                val = b[i]
                                x = b.count(val)
                        indexes = []
                        indexes.append(val)
                        for i in range(len(b)):
                            if b[i] == val:
                                indexes.append(i)

                        return indexes
                    total_aux_list = []
                    aux_list = []
                    print(len(t[0]))
                    try:
                        for team_score_index in range(len(t[0])):
                            for j in range(10):
                                aux_list.append(t[j][team_score_index])
							
                            b = find(aux_list)
                            total_aux_list.append(list(b))
                            aux_list = []
                    except:
                        pass
                    
                    # Append data rows
                    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                        ws.append(list(row))
                    

                    columns_to_check = ['G', 'H', 'I', 'J', 'O', 'P', 'AA', 'AB']

                    # Iterate through rows
                    raw_data = []
                    rr = 1
                    row_index = 0
                    per = []
                    core = ""
                    i = -1
                    k = -1
                    ic(df['AGH']) 
                    ic(df['AGA'])
                    ic(PH)
                    ic(AGA)
                    ic(AGH)
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                        try:
                            for cell in row:
                                i+=1
                                
                                if cell.value is not None and  cell.value != 'HGD' and cell.value=="" and i == 1 and isinstance(cell.value,str):
                                    k+=1
                                    cell.value = str((float(AGH[k]) - float(AGA[k])) * float(PH[k]))
                                    HGD.append(cell.value)
                        except:
                            cell.value = '0'
                            HGD.append(cell.value)
                            i = 0
                            k = 0
                        if i == 11:
                            i = 0     
                    i = -1
                    k = -1
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                        try:
                            for cell in row:
                                i+=1
                                if cell.value is not None and  cell.value != 'AGD' and cell.value=="" and i == 1 and isinstance(cell.value,str):
                                    k+=1
                                    cell.value = str((  float(AGA[k]) - float(AGH[k])) * float(PA[k]))
                                    AGD.append(cell.value)
                                
                        except:
                            cell.value = '0'
                            AGD.append(cell.value)
                            i=0
                            k=0

                        if i == 11:
                            i = 0    

                    ic(HGD)
                    ic(AGD)

                    from itertools import combinations
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=39, max_col=41):
                        for cell in row:
                            try:
                                if cell.value is not None and  cell.value != 'Team' and  cell.value != 'Score' and cell.value != "W/L" and cell.value!="" and isinstance(cell.value,str):
                                    rr+=1
                                    row_index +=1
                                    if cell.value[0] == "(":
                                        core += cell.value[1]
                                    elif cell.value[-2:] == "A)":
                                        core += cell.value[-2]
                                    else:    
                                        core += cell.value
                                if rr %4 == 0:
                                    rr = 1
                                    per.append(core)
                                    core = ""
                                if row_index == 30:
                                    raw_data.append(per)
                                    per = []
                                    row_index = 0
                            except:
                                pass
                    ic(raw_data)

                    kalb = []
                    tot1 = []
                    for team_score_index in range(min(len(GC_home), len(GC_away))):
                        tot1.append([GC_home[team_score_index],GC_away[team_score_index]])
                    tot2 = []
                    for team_score_index in range(min(len(GC_home), len(GC_away))):
                        tot2.append([cv_home[team_score_index],cv_away[team_score_index]])
                    asb = 0
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=20, max_col=20):
                        for cell in row:
                            try:
                                if  cell.value is not None and  cell.value != 'Prob.home' and cell.value !='' and isinstance(cell.value,str) and '.' in cell.value:
                                    som1 = float(tot1[asb][0]) + float(tot1[asb][1])
                                    som2 = float(tot2[asb][0]) + float(tot2[asb][1])
                                    kalb.append(som1-som2)
                                    asb +=1
                                    if abs(som1 - som2) > 0.0 and abs(som1 - som2) <= 0.009:
                                        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                                        font = Font(color='ffffff')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.009 and abs(som1 - som2) <= 0.02:
                                        cell.fill = PatternFill(start_color="1c1c1c", end_color="1c1c1c", fill_type="solid")
                                        font = Font(color='ffffff')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.02 and abs(som1 - som2) <= 0.035:
                                        cell.fill = PatternFill(start_color="393939", end_color="393939", fill_type="solid")
                                        font = Font(color='ffffff')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.035 and abs(som1 - som2) <= 0.05:
                                        cell.fill = PatternFill(start_color="555555", end_color="555555", fill_type="solid")
                                        font = Font(color='ffffff')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.05 and abs(som1 - som2) <= 0.07:
                                        cell.fill = PatternFill(start_color="727272", end_color="727272", fill_type="solid")
                                        font = Font(color='ffffff')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.07 and abs(som1 - som2) <= 0.09:
                                        cell.fill = PatternFill(start_color="8f8f8f", end_color="8f8f8f", fill_type="solid")
                                        font = Font(color='ffffff')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.9 and abs(som1 - som2) <= 0.12:
                                        cell.fill = PatternFill(start_color="acacac", end_color="acacac", fill_type="solid")
                                        font = Font(color='ffffff')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.12 and abs(som1 - som2) < 0.14:
                                        cell.fill = PatternFill(start_color="c9c9c9", end_color="c9c9c9", fill_type="solid")
                                        font = Font(color='000000')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.14 and abs(som1 - som2) < 0.16:
                                        cell.fill = PatternFill(start_color="e6e6e6", end_color="e6e6e6", fill_type="solid")
                                        font = Font(color='000000')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif abs(som1 - som2) > 0.16:
                                        cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
                                        font = Font(color='000000')  # Red color in hexadecimal notation
                                        cell.font = font
                            except:
                                pass  

                    indx = 0
                    las = -1
                    lez = 0
                    le = len(tot1)
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=16, max_col=17):
                        for cell in row:
                            try:
                                if  cell.value is not None and cell.value != 'CV Home' and cell.value != 'CV Away' and cell.value !='' and isinstance(cell.value,str) and '.' in cell.value and cell.value != 'Goal Cost Home' and cell.value != 'Goal Cost Away':
                                    las += 1
                                    if cell.value in tot1[lez] :
                                        som = float(tot1[lez][0]) + float(tot1[lez][1])
                                        if  float(som) <=0.21:
                                            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                                            font = Font(color='4EEA10')  # Red color in hexadecimal notation
                                            cell.font = font
                                        elif float(som) >0.21 and float(som) <=0.39:
                                            cell.fill = PatternFill(start_color="060270", end_color="060270", fill_type="solid")
                                            font = Font(color='FFFFFF')  # Red color in hexadecimal notation
                                            cell.font = font
                                        elif float(som) > 0.39 and float(som) <= 0.51:
                                            cell.fill = PatternFill(start_color="8D26A9", end_color="8D26A9", fill_type="solid")
                                            font = Font(color='FCE95D')  # Red color in hexadecimal notation
                                            cell.font = font
                                        elif float(som) >0.51 and float(som) <= 0.81:
                                            cell.fill = PatternFill(start_color="E4080A", end_color="E4080A", fill_type="solid")
                                            font = Font(color='F3B701')  # Red color in hexadecimal notation
                                            cell.font = font
                                        elif float(som) >0.81 and float(som) <= 1.7:
                                            cell.fill = PatternFill(start_color="E68508", end_color="E68508", fill_type="solid")
                                        elif float(som) > 1.7 and float(som) <= 5.9:
                                            font = Font(color='D20103')  # Red color in hexadecimal notation
                                            cell.font = font
                                    if las == 1:
                                        las = -1
                                        lez += 1
                            except:
                                pass
                    las = -1
                    lez = 0
                    le = len(tot2)
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=18, max_col=19):
                            for cell in row:
                                try:
                                    if  cell.value is not None and cell.value != 'CV Home' and cell.value != 'CV Away' and cell.value !='' and isinstance(cell.value,str) and '.' in cell.value and cell.value != 'Goal Cost Home' and cell.value != 'Goal Cost Away':
                                        las += 1
                                        if cell.value in tot2[lez] :
                                            som = float(tot2[lez][0]) + float(tot2[lez][1])
                                            if  float(som) <=0.1199999:
                                                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                                                font = Font(color='4EEA10')  # Red color in hexadecimal notation
                                                cell.font = font
                                            elif float(som) >0.12 and float(som) <=0.23999999:
                                                cell.fill = PatternFill(start_color="060270", end_color="060270", fill_type="solid")
                                                font = Font(color='FFFFFF')  # Red color in hexadecimal notation
                                                cell.font = font
                                            elif float(som) > 0.24 and float(som) <= 0.759999999:
                                                cell.fill = PatternFill(start_color="8D26A9", end_color="8D26A9", fill_type="solid")
                                                font = Font(color='FCE95D')  # Red color in hexadecimal notation
                                                cell.font = font
                                            elif float(som) >0.76 and float(som) <= 0.9999999:
                                                cell.fill = PatternFill(start_color="E4080A", end_color="E4080A", fill_type="solid")
                                                font = Font(color='F3B701')  # Red color in hexadecimal notation
                                                cell.font = font
                                            elif float(som) >0.9 and float(som) <= 2:
                                                cell.fill = PatternFill(start_color="E68508", end_color="E68508", fill_type="solid")
                                        if las == 1:
                                            las = -1
                                            lez += 1
                                except:
                                    pass
                    except:
                        pass
                    new_avg = []


                    t1 = last_matches1
                    t2 = last_matches2
                    t3 = []
                    try:
                        for team_score_index in range(0,len(t1),2):
                            t3.append([t1[team_score_index][0],t1[team_score_index][1],t1[team_score_index+1][0],t1[team_score_index+1][1],t2[team_score_index][0],t2[team_score_index][1],t2[team_score_index+1][0],t2[team_score_index+1][1]])                
                    except:
                        pass

                    def null_to_null(t):
                        res = []
                        res2 = []
                        try:
                            for i in range(len(t)):
                                for j in range(0,len(t[i]),2):
                                    if t[i][j] + t[i][j+1] == 0:
                                        res.append(True)
                                        
                                    else:
                                        res.append(False)
                                res2.append(res)
                                res = []
                        except:
                            pass

                        return res2

                    start_cell = 32
                    end_cell = 33
                    ic(t3)
                    rr = 1
                    row_index = 0
                    results = []
                    per = []
                    core = ""
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=40, max_col=41):
                        for cell in row:
                            try:
                                if cell.value is not None and  cell.value != 'Score' and cell.value != "W/L" and cell.value!="" and isinstance(cell.value,str):
                                    rr+=1
                                    row_index +=1
                                    core += cell.value
                                if rr %3 == 0:
                                    rr = 1
                                    per.append(core)
                                    core = ""
                                if row_index == 20:
                                    results.append(per)
                                    per = []
                                    row_index = 0
                            except:
                                pass

                    def three_in_rowH(wl,t):
                        t = [row[:5] for row in t]
                        ic(t)
                        found = 0
                        data =[]

                        for i in range(len(t)):
                            found = 0
                            for j in range(len(t[i])-2):         
                                if t[i][j][-1] == wl and t[i][j+1][-1] == wl and t[i][j+2][-1] == wl :
                                    found = 1
                            if found == 1:
                                data.append(True)
                            else:
                                data.append(False)

                        return data
                    checkW = three_in_rowH("W",results)
                    checkD = three_in_rowH("D",results)
                    checkL = three_in_rowH("L",results)
                    rw = -1
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6, max_col=6):
                        for cell in row:
                            try:
                                if cell.value is not None and  cell.value != 'GH' and cell.value!="" and isinstance(cell.value,str):
                                    rw += 1
                                    if checkW[rw] == True:
                                        cell.fill = PatternFill(start_color="5BFA1C", end_color="5BFA1C", fill_type="solid")
                                    if checkD[rw] == True:
                                        cell.fill = PatternFill(start_color="FDF10B", end_color="FDF10B", fill_type="solid")
                                    if checkL[rw] == True and PH[rw]<=PA[rw]:
                                        cell.fill = PatternFill(start_color="FD0B0B", end_color="FD0B0B", fill_type="solid")
                            except:
                                pass        
                    def three_in_rowA(wl,t):
                        found = 0
                        data =[]
                        ic(t)
                        t = [row[5:] for row in t]

                        for i in range(len(t)):
                            found = 0
                            for j in range(len(t[i])-2):         
                                if t[i][j][-1] == wl and t[i][j+1][-1] == wl and t[i][j+2][-1] == wl :
                                    found = 1
                            if found == 1:
                                data.append(True)
                            else:
                                data.append(False)

                        return data
                    checkW = three_in_rowA("W",results)
                    checkD = three_in_rowA("D",results)
                    checkL = three_in_rowA("L",results)
                    rw = -1
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=7, max_col=7):
                        for cell in row:
                            try:
                                if cell.value is not None and  cell.value != 'GA' and cell.value!="" and isinstance(cell.value,str):
                                    rw += 1
                                    if checkW[rw] == True:
                                        cell.fill = PatternFill(start_color="5BFA1C", end_color="5BFA1C", fill_type="solid")
                                    if checkD[rw] == True:
                                        cell.fill = PatternFill(start_color="FDF10B", end_color="FDF10B", fill_type="solid")
                                    if checkL[rw] == True and PH[rw]>=PA[rw]:
                                        cell.fill = PatternFill(start_color="FD0B0B", end_color="FD0B0B", fill_type="solid")
                            except:
                                pass 
                    print("Verifying data...")
                    home =[row[0:5] for row in raw_data]
                    away =[row[5:] for row in raw_data]      
                    ic(home)
                    ic(away)                             
                    def verify_data(home,away):
                        for elem in home:
                            for i in range(len(elem)):
                                if len(elem[i]) <10:
                                    elem[i] = 'A3-0(1-0)L'
                        for elem in away:
                            for i in range(len(elem)):
                                if len(elem[i]) <10:
                                    elem[i] = 'A3-0(1-0)L'
                        return home,away

                    home,away = verify_data(home,away)
                    ################################################################ D3rd (Modified) ##########################################################################
                    print('Verified succsesfuly')
                    def new_task1(home,away):
                        found = False
                        resH = []
                        for elem in home:
                            found = False
                            for i in range(len(elem)):
                                if i == 2 and elem[i][1] == '2' and elem[i][3] == '2':
                                    found = True
                                elif i == 2 and elem[i][1] == '3' and elem[i][3] == '3':
                                    found = True
                            if found:
                                resH.append(True)
                            else:
                                resH.append(False)
                        found = False
                        resA = []
                        for elem in away:
                            found = False
                            for i in range(len(elem)):
                                if i == 2 and elem[i][1] == '2' and elem[i][3] == '2':
                                    found = True
                                elif i == 2 and elem[i][1] == '3' and elem[i][3] == '3':
                                    found = True
                            if found:
                                resA.append(True)
                            else:
                                resA.append(False)
                                
                        return resH, resA
                    font
                    resH, resA = new_task1(home,away)
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6, max_col=6):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'GH' and cell.value!="" and isinstance(cell.value,str):
                                    i+= 1
                                    r = float(SDH[i]) + float(SDA[i])
                                    if resH[i] == True and float(r) <0.54:
                                        font = Font(color='0000FF',name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                    elif resH[i] == True and resA[i] == True and float(r) <0.54:
                                        font = Font(color='FF0000',name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                    PROV_D3RD = True
                                    r = 0
                    except:
                        pass

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=7, max_col=7):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'GA' and cell.value!="" and isinstance(cell.value,str):
                                    i+= 1
                                    if resA[i] == True:
                                        font = Font(color='0000FF',name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                    elif resH[i] == True and resA[i] == True:
                                        font = Font(color='FF0000',name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                    PROV_D3RD = True
                    except:
                        pass

                    ################################################################ D5th (Modified) ##########################################################################
                    def new_task_update_1(home,away):
                        found = False
                        resH = []
                        for elem in home:
                            found = False
                            for i in range(len(elem)):
                                
                                if i == 4 and elem[i][1] == '2' and elem[i][3] == '2':
                                    found = True
                                elif i == 4 and elem[i][1] == '3' and elem[i][3] == '3':
                                    found = True
                            if found:
                                resH.append(True)
                            else:
                                resH.append(False)
                        found = False
                        resA = []
                        for elem in away:
                            found = False
                            for i in range(len(elem)):
                                if i == 4 and elem[i][1] == '2' and elem[i][3] == '2':
                                    found = True
                                elif i == 4 and elem[i][1] == '3' and elem[i][3] == '3':
                                    found = True
                            if found:
                                resA.append(True)
                            else:
                                resA.append(False)

                        return resH, resA
                    font
                    resH, resA = new_task_update_1(home,away)

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                            for cell in row:
                                if cell.value is not None and cell.value!="" and cell.value!="Time" and isinstance(cell.value,str):
                                    i+= 1
                                    if resH[i] == True:
                                        cell.fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
                                        #font = Font(color='#92CDDC',name='Arial Narrow', size=11, bold=True, italic=True)
                                        #cell.font = font
                                        PROV_D5TH = True
                    except:
                        pass
                    ################################################################ HDN (Modified) ##########################################################################
                    def new_task_update_2(home, away):
                        resH = []
                        resA = []
                        # Verifica a lista "home"
                        for elem in home:
                            found = False
                            for i in range(len(elem)):
                                
                                if i == 4:
                                    if (elem[i][1] == '0' and elem[i][3] == '0') or (elem[i][1] == '1' and elem[i][3] == '1'):
                                        found = True
                            resH.append(found)

                        # Verifica a lista "away"
                        for elem in away:
                            found = False
                            for i in range(len(elem)):
                                if i == 4:
                                    if (elem[i][1] == '0' and elem[i][3] == '0') or (elem[i][1] == '1' and elem[i][3] == '1'):
                                        found = True
                            resA.append(found)

                        return resH, resA
                    resH, resA = new_task_update_2(home, away)

                    i = -1

                    odds: list = []
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=24, max_col=24):
                        for cell in row:
                            if cell.value is not None and cell.value != "" and cell.value != "Early" and isinstance(cell.value, str):
                                odds.append(cell.value)

                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=9, max_col=9):
                            for cell in row:
                                if cell.value is not None and cell.value != "" and cell.value != "AVGA" and isinstance(cell.value, str):
                                    i += 1
                                    if i < len(resH) and resH[i] == True:
                                        if (float(odds[i].split('/')[0].strip()) >= float(odds[i].split('/')[2].strip())):
                                            font = Font(color='FF0000FF', name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            PROV_HDN = True
                                        else:
                                            font = Font(color='FFFACA00', name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            PROV_HDN = True
                                    elif i < len(resA) and resA[i] == True:
                                        if (float(odds[i].split('/')[0].strip()) <= float(odds[i].split('/')[2].strip())):
                                            font = Font(color='FF0000FF', name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            PROV_HDN = True
                                        else:
                                            font = Font(color='FFFACA00', name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            PROV_HDN = True
                                    elif((i < len(resA) and resA[i] == True) and (i < len(resH) and resH[i] == True)):
                                        if (float(odds[i].split('/')[0].strip()) >= float(odds[i].split('/')[2].strip())):
                                            font = Font(color='FF0000FF', name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            PROV_HDN = True
                                        else:
                                            font = Font(color='FFFACA00', name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            PROV_HDN = True
                                    elif i < len(resA) and resA[i] == True:
                                        if (float(odds[i].split('/')[0].strip()) <= float(odds[i].split('/')[2].strip())):
                                            font = Font(color='FF0000FF', name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            PROV_HDN = True
                                        else:
                                            font = Font(color='FFFACA00', name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            PROV_HDN = True
                    except Exception as e:
                        print(f"Erro ao processar a planilha na célula {cell.coordinate}: {e}")
                    
                   ################################################################ DIR (Modified) ##########################################################################
                    def new_task_update_3(home,away):
                        resH = []
                        resA = []
                        # Verifica a lista "home"
                        for elem in home:
                            found = False
                            for i in range(len(elem)):
                                
                                if i == 0 and ('W' == elem[i][-1] or 'D' == elem[i][-1] or 'w' == elem[i][-1] or 'd' == elem[i][-1]):
                                    if (f'{elem[i][1]}-{elem[i][3]}' in ['2-2', '3-3', '2-3', '1-3', '2-4', '0-4', '1-4', '3-4', '1-5', '0-5', '2-5', '3-5']):
                                        found = True
                            resH.append(found)  # Adiciona True ou False para cada elemento

                        # Verifica a lista "away"
                        for elem in away:
                            found = False
                            for i in range(len(elem)):
                                # Verifica as condições no índice 4
                                if i == 0 and ('W' == elem[i][-1] or 'D' == elem[i][-1] or 'w' == elem[i][-1] or 'd' == elem[i][-1]):
                                    if (f'{elem[i][1]}-{elem[i][3]}' in ['2-2', '3-3', '3-1', '3-2', '4-0', '4-1', '4-2', '4-3', '5-0', '5-1', '5-2', '5-3']):
                                        found = True
                            resA.append(found)  # Adiciona True ou False para cada elemento

                        return resH, resA
                    resH, resA = new_task_update_3(home, away)

                    i = -1  # Inicializa o índice

                    try:
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):  # Itera na coluna 9
                            for cell in row:
                                # Verifica se a célula contém um valor válido
                                if cell.value is not None and cell.value != "" and cell.value != "AVGH" and isinstance(cell.value, str):
                                    i += 1
                                    if i < len(resH) and resH[i] == True:
                                        font = Font(color='FFFFFF00', name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                        PROV_D1R = True
                                    elif i < len(resA) and resA[i] == True:
                                        font = Font(color='FFFFFF00', name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                        PROV_D1R = True
                    except Exception as e:
                        print(f"Erro ao processar a planilha na célula {cell.coordinate}: {e}")
                    ################################################################ DIR (Modified) ##########################################################################
                    def new_task_update_4(home, away):
                        found = False
                        lis_color = [[], []]  # Lista de cores separada para home e away
                        resH = []  # Resultado para a lista "home"
                        resA = []  # Resultado para a lista "away"

                        # Verifica a lista "home"
                        for elem in home:
                            found = False
                            for i in range(len(elem)):
                                if i == 4 and (int(elem[i][1]) + int(elem[i][3])) >= 5:
                                    found = True
                                    if elem[i][-1].upper() == 'W' or elem[i][-1].upper() == 'D':
                                        lis_color[0].append('FF0000FF')  # Azul
                                    elif elem[i][-1].upper() == 'L':
                                        lis_color[0].append('FFFF0000')  # Vermelho
                            resH.append(found)

                        # Verifica a lista "away"
                        for elem in away:
                            found = False
                            for i in range(len(elem)):
                                if i == 4 and (int(elem[i][1]) + int(elem[i][3])) >= 5:
                                    found = True
                                    if elem[i][-1].upper() == 'W' or elem[i][-1].upper() == 'D':
                                        lis_color[1].append('FF0000FF')  # Azul
                                    elif elem[i][-1].upper() == 'L':
                                        lis_color[1].append('FFFF0000')  # Vermelho
                            resA.append(found)

                        return resH, resA, lis_color


                    # Obtém os resultados e as cores
                    resH, resA, color_d77 = new_task_update_4(home, away)

                    indexH = -1
                    indexA = -1 
                    i = -1  # Índice global para as linhas da planilha

                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):  # Apenas coluna 3
                            for cell in row:
                                # Verifica se a célula contém um valor válido
                                if cell.value is not None and cell.value != "" and cell.value != "League" and isinstance(cell.value, str):
                                    i += 1
                                    if i < len(resH) and resH[i] == True:
                                        indexH += 1  # Incrementa o índice para color_d77[0]
                                        border_color = Border(
                                            left=Side(border_style='thick', color=color_d77[0][indexH]),
                                            right=Side(border_style='thick', color=color_d77[0][indexH]),
                                            top=Side(border_style='thick', color=color_d77[0][indexH]),
                                            bottom=Side(border_style='thick', color=color_d77[0][indexH]),
                                        )
                                        cell.border = border_color  # Aplica borda
                                        PROV_D77 = True
                                    elif i < len(resA) and resA[i] == True:
                                        indexA += 1  # Incrementa o índice para color_d77[1]
                                        border_color = Border(
                                            left=Side(border_style='thick', color=color_d77[1][indexA]),
                                            right=Side(border_style='thick', color=color_d77[1][indexA]),
                                            top=Side(border_style='thick', color=color_d77[1][indexA]),
                                            bottom=Side(border_style='thick', color=color_d77[1][indexA]),
                                        )
                                        cell.border = border_color  # Aplica borda
                                        PROV_D77 = True

                    except Exception as e:
                        print(f"Erro ao processar a planilha na célula {cell.coordinate}: {e}")
                    ################################################################ DIR (Modified) ##########################################################################
                    def new_task_update_5(home,away):
                        found = False
                        resH = []
                        for elem in home:
                            found = False
                            for i in range(len(elem)):
                                if i == 4 and (elem[i][1] == '4' or elem[i][3] == '5'):
                                    found = True
                            if found:
                                resH.append(True)
                            else:
                                resH.append(False)
                        found = False
                        resA = []
                        for elem in away:
                            found = False
                            for i in range(len(elem)):
                                if i == 4 and (elem[i][1] == '4' or elem[i][3] == '5'):
                                    found = True
                            if found:
                                resA.append(True)
                            else:
                                resA.append(False)

                        return resH, resA
                    font
                    resH, resA = new_task_update_5(home,away)

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):
                            for cell in row:
                                if cell.value is not None and cell.value!="" and isinstance(cell.value,str):
                                    i+= 1
                                    if resH[i] == True:
                                        cell.border = color111
                    except:
                        pass

                    def new_task2(home,away):
                        found = False
                        resH = []
                        for elem in home:
                            found = False
                            for i in range(len(elem)):
                                
                                if i == 4 and elem[i][1] == '2' and elem[i][3] == '2':
                                    found = True
                                elif i == 4 and elem[i][1] == '3' and elem[i][3] == '3':
                                    found = True
                            if found:
                                resH.append(True)
                            else:
                                resH.append(False)
                        found = False
                        resA = []
                        for elem in away:
                            found = False
                            for i in range(len(elem)):
                                
                                if i == 4 and elem[i][1] == '2' and elem[i][3] == '2':
                                    found = True
                                elif i == 4 and elem[i][1] == '3' and elem[i][3] == '3':
                                    found = True
                            if found:
                                resA.append(True)
                            else:
                                resA.append(False)

                        return resH, resA
                    ################################################################ NEW TASK 2 ##########################################################################
                    resHH, resAA = new_task2(home,away)
                    ic(resHH)
                    ic(resAA)
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'Time' and cell.value!="" and isinstance(cell.value,str):
                                    i+= 1
                                    if resHH[i] == True:
                                        cell.fill = PatternFill(start_color="5DE2E7", end_color="5DE2E7", fill_type="solid")
                                    elif resAA[i] == True:
                                        cell.fill = PatternFill(start_color="5DE2E7", end_color="5DE2E7", fill_type="solid")
                    except:
                        pass

                    ################################# ECC ###########################
                    ECC = []
                    i = -1 
                    isnegative = False
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'HGD' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if cell.value[0] == '-':
                                        isnegative = True
                                        cell.value = cell.value[1:] 
                                    if len(cell.value) > 3 and len(CVH[i]) > 3 and len(CVA[i]) > 3 and (cell.value[cell.value.index('.')+1:4] == CVH[i][CVH[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == CVA[i][CVA[i].index('.')+1:4]  or cell.value[cell.value.index('.')+1:4] == GCH[i][GCH[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == GCA[i][GCA[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == PH[i][PH[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == PA[i][PA[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == SDH[i][SDH[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == SDA[i][SDA[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == CVA[i][CVA[i].index('.')+1:4]):
                                        font = Font(color='ffa500',name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                        ECC.append(True)
                                        PROV_ECC = True
                                    else:
                                        ECC.append(True) # Should be False
                                        PROV_ECC = True
                                if isnegative == True:
                                    cell.value = '-'+cell.value
                                    isnegative = False
                    except:
                        pass

                    i = -1 
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'AGD' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if cell.value[0] == '-':
                                        isnegative = True
                                        cell.value = cell.value[1:]
                                    if len(cell.value) > 3 and len(CVH[i]) > 3 and len(CVA[i]) > 3 and (cell.value[cell.value.index('.')+1:4] == CVH[i][CVH[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == CVA[i][CVA[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == GCH[i][GCH[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == GCA[i][GCA[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == PH[i][PH[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == PA[i][PA[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == SDH[i][SDH[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == SDA[i][SDA[i].index('.')+1:4] or cell.value[cell.value.index('.')+1:4] == CVA[i][CVA[i].index('.')+1:4]):
                                        font = Font(color='ffa500',name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                        ECC[i] = True
                                        PROV_ECC = True
                                    else:
                                        ECC[i] = False
                            if isnegative == True:
                                    cell.value = '-'+cell.value
                                    isnegative = False
                    except:
                        pass

                    ################################################################ Zoya ##########################################################################
                    ZOYA = []
                    ZOYAF_09 = []
                    Pink_Zoya = []
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=24, max_col=24):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'Early' and cell.value!="" and isinstance(cell.value,str):
                                    i+=1
                                    
                                    odd = EARLY[i][EARLY[i].index('/')+1:]
                                    odd = odd[:EARLY[i].index('/')]
                                    ic(odd)
                                    odd = float(odd)
                                    ic(odd)
                                    if float(PH[i]) >= 0.6369426751592357  and float(PH[i]) <=  0.9900990099009901 and float(AVGH[i]) > float(AVGA[i]) and float(AGH[i]) > float(AGA[i]) and float(SDH[i]) < float(SDA[i]):
                                        cell.fill = PatternFill(start_color="FF7D7D", end_color="FF7D7D", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(True)
                                        Pink_Zoya.append(False)
                                    elif float(PH[i]) >= 0.3802281368821293  and float(PH[i]) <=  0.9900990099009901 and float(AVGH[i]) > float(AVGA[i]) and float(AGH[i]) < float(AGA[i]) and float(SDH[i]) > float(SDA[i]):
                                        cell.fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(False)
                                        Pink_Zoya.append(True)
                                    elif float(PH[i]) >= 0.3802281368821293  and float(PH[i]) <=  0.9900990099009901 and float(AVGH[i]) < float(AVGA[i]) and float(AGH[i]) > float(AGA[i]) and float(SDH[i]) > float(SDA[i]):
                                        cell.fill = PatternFill(start_color="85DFFF", end_color="85DFFF", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(False)
                                        Pink_Zoya.append(True)
                                    elif float(PH[i]) >= 0.3802281368821293  and float(PH[i]) <=  0.9900990099009901 and float(AVGH[i]) < float(AVGA[i]) and float(AGH[i]) < float(AGA[i]) and float(SDH[i]) > float(SDA[i]) and ((float(SDH[i])+float(SDA[i])) <= 0.57 and (odd >= 3.40)or (ECC[i] == True)):
                                        cell.fill = PatternFill(start_color="E5B577", end_color="E5B577", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(False)
                                        Pink_Zoya.append(True)
                                    elif float(PH[i]) >= 0.3802281368821293  and float(PH[i]) <=  0.625 and float(AVGH[i]) > float(AVGA[i]) and float(AGH[i]) > float(AGA[i]) and float(SDH[i]) < float(SDA[i]) and ((float(SDH[i])+float(SDA[i])) <= 0.57 and (odd >= 3.40)or (ECC[i] == True)):
                                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(True)
                                        Pink_Zoya.append(False)
                                    elif float(PH[i]) >= 0.3802281368821293  and float(PH[i]) <=  0.9900990099009901 and float(AVGH[i]) < float(AVGA[i]) and float(AGH[i]) > float(AGA[i]) and float(SDH[i]) < float(SDA[i]) and (float(SDH[i])+float(SDA[i])) <= 0.81:
                                        cell.fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(True)
                                        Pink_Zoya.append(False)
                                    elif float(AGD[i]) == 0  and float(PH[i]) > float(PA[i]) and float(AVGH[i]) < float(AVGA[i]) and float(SDH[i]) > float(SDA[i]):
                                        ZOYA.append(True)
                                        ZOYAF_09.append(False)
                                        Pink_Zoya.append(False)
                                        cell.fill = PatternFill(start_color="CCFF66", end_color="CCFF66", fill_type="solid")
                                    elif float(AGD[i]) == 0  and float(PH[i]) > float(PA[i]) and float(AVGH[i]) > float(AVGA[i]) and float(SDH[i]) < float(SDA[i]):
                                        ZOYA.append(True)
                                        ZOYAF_09.append(False)
                                        Pink_Zoya.append(False)
                                        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                                    elif float(PH[i]) >= 0.3802281368821293 and float(PH[i]) <= 0.625  and float(AVGH[i]) > float(AVGA[i]) and float(AGH[i]) < float(AGA[i]) and  float(SDH[i]) < float(SDA[i]) and (float(SDH[i])+float(SDA[i])) <= 0.71 and (odd >= 3.40):
                                        cell.fill = PatternFill(start_color="403151", end_color="403151", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(True)
                                        Pink_Zoya.append(False)
                                    elif  float(AVGH[i]) < float(AVGA[i]) and float(AGH[i]) <= float(AGA[i]) and  float(SDH[i]) < float(SDA[i]) and (float(SDH[i])+float(SDA[i])) <= 0.71 :
                                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(True)
                                        Pink_Zoya.append(False)
                                    elif  float(AVGH[i]) >= float(AVGA[i]) and float(AGH[i]) > float(AGA[i]) and  float(SDH[i]) > float(SDA[i]) and (float(SDH[i])+float(SDA[i])) <= 0.71 :
                                        cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
                                        ZOYA.append(True)
                                        ZOYAF_09.append(False)
                                        Pink_Zoya.append(True)
                                    else:
                                        ZOYA.append(False)    
                                        ZOYAF_09.append(False)
                                        Pink_Zoya.append(False)
                    except:
                        pass 

                    ZOYA2 = []
                    Pink_Zoya2 = []
                    i = -1      
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=24, max_col=24):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'Early' and cell.value!="" and isinstance(cell.value,str):
                                    i+=1
                                    odd = EARLY[i][EARLY[i].index('/')+1:]
                                    odd = odd[:EARLY[i].index('/')]
                                    ic(odd)
                                    odd = float(odd)
                                    ic(odd)
                                    if float(PA[i]) >= 0.6369426751592357  and float(PA[i]) <=  0.9900990099009901 and float(AVGA[i]) > float(AVGH[i]) and float(AGA[i]) > float(AGH[i]) and float(SDA[i]) < float(SDH[i]):
                                        cell.fill = PatternFill(start_color="FF7D7D", end_color="FF7D7D", fill_type="solid")
                                        ZOYA2.append(True)
                                        Pink_Zoya2.append(False)
                                        ZOYAF_09[i] = True
                                    elif float(PA[i]) >= 0.3802281368821293  and float(PA[i]) <=  0.9900990099009901 and float(AVGA[i]) > float(AVGH[i]) and float(AGA[i]) < float(AGH[i]) and float(SDA[i]) > float(SDH[i]):
                                        cell.fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
                                        ZOYA2.append(True)
                                        Pink_Zoya2.append(True)
                                        ZOYAF_09[i] = False
                                    elif float(PA[i]) >= 0.3802281368821293  and float(PA[i]) <=  0.9900990099009901 and float(AVGA[i]) < float(AVGH[i]) and float(AGA[i]) > float(AGH[i]) and float(SDA[i]) > float(SDH[i]):
                                        cell.fill = PatternFill(start_color="85DFFF", end_color="85DFFF", fill_type="solid")
                                        ZOYA2.append(True)
                                        Pink_Zoya2.append(True)
                                        ZOYAF_09[i] = False
                                    elif float(PA[i]) >= 0.3802281368821293  and float(PA[i]) <=  0.9900990099009901 and float(AVGA[i]) < float(AVGH[i]) and float(AGA[i]) < float(AGH[i]) and float(SDA[i]) > float(SDH[i]) and ((float(SDH[i])+float(SDA[i])) <= 0.57 and (odd >= 3.40)or (ECC[i] == True)):
                                        cell.fill = PatternFill(start_color="E5B577", end_color="E5B577", fill_type="solid")
                                        ZOYA2.append(True)
                                        ZOYAF_09[i] = False
                                        Pink_Zoya2.append(True)
                                    elif float(PA[i]) >= 0.3802281368821293  and float(PA[i]) <=  0.625 and float(AVGA[i]) > float(AVGH[i]) and float(AGA[i]) > float(AGH[i]) and float(SDA[i]) < float(SDH[i]) and ((float(SDH[i])+float(SDA[i])) <= 0.57 and (odd >= 3.40) or ECC[i] == True ):
                                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                                        ZOYA2.append(True)
                                        Pink_Zoya2.append(False)
                                        ZOYAF_09[i] = True
                                    elif float(PA[i]) >= 0.3802281368821293  and float(PA[i]) <=  0.9900990099009901 and float(AVGH[i]) < float(AVGA[i]) and float(AGH[i]) > float(AGA[i]) and float(SDH[i]) < float(SDA[i]):
                                        cell.fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")
                                        ZOYA2.append(True)
                                        ZOYAF_09[i] = True
                                        Pink_Zoya2.append(False)
                                    elif float(AGD[i]) == 0  and float(PH[i]) < float(PA[i]) and float(AVGH[i]) > float(AVGA[i]) and float(SDH[i]) < float(SDA[i]) and (float(SDH[i])+float(SDA[i])) <= 0.81:
                                        cell.fill = PatternFill(start_color="CCFF66", end_color="CCFF66", fill_type="solid")
                                        ZOYA2.append(True)
                                        ZOYAF_09[i] = False
                                        Pink_Zoya2.append(False)
                                    elif float(AGD[i]) == 0  and float(PH[i]) < float(PA[i]) and float(AVGH[i]) < float(AVGA[i]) and float(SDH[i]) > float(SDA[i]):
                                        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                                        ZOYA2.append(True)
                                        ZOYAF_09[i] = False
                                        Pink_Zoya2.append(False)
                                    elif float(PA[i]) >= 0.3802281368821293 and float(PA[i]) <= 0.625  and float(AVGH[i]) < float(AVGA[i]) and float(SDH[i]) > float(SDA[i]) and float(AGH[i]) > float(AGA[i]) and  (float(SDH[i])+float(SDA[i])) <= 0.71 and (odd >= 3.40):
                                        cell.fill = PatternFill(start_color="403151", end_color="403151", fill_type="solid")
                                        ZOYA2.append(True)
                                        Pink_Zoya2.append(False)
                                        ZOYAF_09[i] = True
                                    elif  float(AVGH[i]) > float(AVGA[i]) and float(AGH[i]) >= float(AGA[i]) and  float(SDH[i]) > float(SDA[i]) and (float(SDH[i])+float(SDA[i])) <= 0.71 :
                                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                        ZOYA2.append(True)
                                        ZOYAF_09[i] = True
                                        Pink_Zoya2.append(False)
                                    elif  float(AVGH[i]) <= float(AVGA[i]) and float(AGH[i]) < float(AGA[i]) and  float(SDH[i]) < float(SDA[i]) and (float(SDH[i])+float(SDA[i])) <= 0.71 :
                                        cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
                                        ZOYA2.append(True)
                                        ZOYAF_09[i] = False
                                        Pink_Zoya2.append(True)
                                    else:
                                        Pink_Zoya2.append(False)
                                        ZOYA2.append(False)
                    except:
                        pass   
                    #GH
                    home =[row[0:5] for row in raw_data]
                    away =[row[5:] for row in raw_data]      
                    ic(home)
                    ic(away)                  
                    
                    ########################################################## NULL TO NULL ###############################################
                    def new_null_to_null(home,away):
                        home_home = []
                        home_away = []
                        for i in range (len(home)):
                            for j in range(len(home[i])):
                                if home[i][j][0] == "H":
                                    home_home.append(home[i][j])
                                    break
                        for i in range (len(home)):
                            for j in range(len(home[i])):
                                if home[i][j][0] == "A":
                                    home_away.append(home[i][j])
                                    break
                        if len(home_home) == len(home_away):
                            away_home = []
                            away_away = []

                            for i in range (len(away)):
                                for j in range(len(away[i])):
                                    if away[i][j][0] == "H":
                                        away_home.append(away[i][j])
                                        break
                            for i in range (len(away)):
                                for j in range(len(away[i])):
                                    if away[i][j][0] == "A":
                                        away_away.append(away[i][j])
                                        break
                            data = []
                            if len(away_home) == len(away_away):
                                for i in range (len(home_home)):
                                    data.append([home_home[i],home_away[i],away_home[i],away_away[i]])
                            else:
                                return None
                            
                            res = []
                            for i in  range (len(data)):
                                x = 0
                                if data[i][0][-1] == "D" and data[i][0][1] == "0" and data[i][0][3] == "0":
                                    x +=1
                                elif data[i][1][-1] == "D" and data[i][1][1] == "0" and data[i][1][3] == "0":
                                    x +=2
                                elif data[i][2][-1] == "D" and data[i][2][1] == "0" and data[i][2][3] == "0":
                                    x +=3
                                elif data[i][3][-1] == "D" and data[i][3][1] == "0" and data[i][3][3] == "0":
                                    x +=4
                                if x == 1 :
                                    res.append([True,False])
                                elif x == 2:
                                    res.append([True,False])
                                elif x == 3 :
                                    res.append([False,True])
                                elif x == 4 :
                                    res.append([False,True])
                                elif x > 6:
                                    res.append([True,True])
                                else:
                                    res.append([False,False])
                            return res
                        else:
                            return None
                    res = new_null_to_null(home,away)
                    ic(res)
                    ic(home)
                    ic(away)
                    ################################################################ LN-7 TASK ##############################################
                    def LN_7(home,away):
                        no = ['1-0', '0-1', '2-0', '0-2']
                        yes = ['1-5','5-1', '4-2','2-4', '6-0','0-6']
                        resH = []
                        for row in home:
                            some = 0
                            for i in range (len(row)):
                                if row[i][-1] == 'L' and row[i][1:4] not in no and row[i][1:4] not in yes:
                                    some += int(row[i][1]) + int(row[i][3])
                                elif row[i][-1] == 'L' and row[i][1:4] in yes:
                                    some = 7
                            if some >= 7:
                                resH.append(True)
                            else:
                                resH.append(False)

                        ################## AWAY ###############
                        resA = []
                        for row in away:
                            some = 0
                            for i in range (len(row)):
                                if row[i][-1] == 'L' and row[i][1:4] not in no and row[i][1:4] not in yes:
                                    some += int(row[i][1]) + int(row[i][3])
                                elif row[i][-1] == 'L' and row[i][1:4] in yes:
                                    some = 7
                            if some >= 7:
                                resA.append(True)
                            else:
                                resA.append(False)
                        return resH,resA

                    Res_Home, Res_Away = LN_7(home,away)

                    thick_red_border = Border(left=Side(border_style='thick', color='FF0000'),
                          right=Side(border_style='thick', color='FF0000'),
                          top=Side(border_style='thick', color='FF0000'),
                          bottom=Side(border_style='thick', color='FF0000'))
                    thick_black_border = Border(left=Side(border_style='thick', color='000000'),
                          right=Side(border_style='thick', color='000000'),
                          top=Side(border_style='thick', color='000000'),
                          bottom=Side(border_style='thick', color='000000'))
                    color1 = Border(left=Side(border_style='thick', color='00B0F0'),
                          right=Side(border_style='thick', color='00B0F0'),
                          top=Side(border_style='thick', color='00B0F0'),
                          bottom=Side(border_style='thick', color='00B0F0'))
                    color2 = Border(left=Side(border_style='thick', color='FF00FF'),
                          right=Side(border_style='thick', color='FF00FF'),
                          top=Side(border_style='thick', color='FF00FF'),
                          bottom=Side(border_style='thick', color='FF00FF'))
                    color3 = Border(left=Side(border_style='thick', color='FF0000'),
                          right=Side(border_style='thick', color='FF0000'),
                          top=Side(border_style='thick', color='FF0000'),
                          bottom=Side(border_style='thick', color='FF0000'))
                    color4 = Border(left=Side(border_style='thick', color='FFC000'),
                          right=Side(border_style='thick', color='FFC000'),
                          top=Side(border_style='thick', color='FFC000'),
                          bottom=Side(border_style='thick', color='FFC000'))
                    color5 = Border(left=Side(border_style='thick', color='00FF00'),
                          right=Side(border_style='thick', color='00FF00'),
                          top=Side(border_style='thick', color='00FF00'),
                          bottom=Side(border_style='thick', color='00FF00'))
                    color6 = Border(left=Side(border_style='thick', color='FFFF00'),
                          right=Side(border_style='thick', color='FFFF00'),
                          top=Side(border_style='thick', color='FFFF00'),
                          bottom=Side(border_style='thick', color='FFFF00'))
                    color7 = Border(left=Side(border_style='thick', color='00FFFF'),
                          right=Side(border_style='thick', color='00FFFF'),
                          top=Side(border_style='thick', color='00FFFF'),
                          bottom=Side(border_style='thick', color='00FFFF'))
                    color8 = Border(left=Side(border_style='thick', color='FF0000'),
                          right=Side(border_style='thick', color='FF0000'),
                          top=Side(border_style='thick', color='FF0000'),
                          bottom=Side(border_style='thick', color='FF0000'))
                    color9 = Border(left=Side(border_style='thick', color='00FF00'),
                          right=Side(border_style='thick', color='00FF00'),
                          top=Side(border_style='thick', color='00FF00'),
                          bottom=Side(border_style='thick', color='00FF00'))
                    color10 = Border(left=Side(border_style='thick', color='FF00FF'),
                          right=Side(border_style='thick', color='FF00FF'),
                          top=Side(border_style='thick', color='FF00FF'),
                          bottom=Side(border_style='thick', color='FF00FF'))
                    color11 = Border(left=Side(border_style='thick', color='0000FF'),
                          right=Side(border_style='thick', color='0000FF'),
                          top=Side(border_style='thick', color='0000FF'),
                          bottom=Side(border_style='thick', color='0000FF'))
                    SCA_Check = False
                    
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=10, max_col=10):
                            for cell in row:
                                    if cell.value is not None and  cell.value != 'AGH' and cell.value !='' and isinstance(cell.value,str) :
                                        i += 1
                                        if Res_Home[i] == True:
                                            font = Font(color='FF0000',name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            if AVGH[i] < AVGA[i] and AGH[i] < AGA[i]:
                                                cell.fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
                                            elif PH[i] >= PA[i] and AVGH[i] >= AVGA[i] and AGH[i] < AGA[i]:
                                                cell.fill = PatternFill(start_color="CCFF66", end_color="CCFF66", fill_type="solid")
                                            elif PH[i] >= PA[i] and AVGH[i] >= AVGA[i] and AGH[i] > AGA[i]:
                                                cell.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
                                            elif PH[i] <= PA[i] and AVGH[i] >= AVGA[i] and AGH[i] >= AGA[i]:
                                                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                                            elif PH[i] <= PA[i] and AVGH[i] >= AVGA[i] and AGH[i] < AGA[i]:
                                                cell.fill = PatternFill(start_color="B1A0C7", end_color="B1A0C7", fill_type="solid")
                    except:
                        pass
                                        
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=11, max_col=11):
                            for cell in row:
                                    if cell.value is not None and  cell.value != 'AGA' and cell.value !='' and isinstance(cell.value,str) :
                                        i += 1
                                        if Res_Away[i] == True:
                                            font = Font(color='FF0000',name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                            if AVGH[i] > AVGA[i] and AGH[i] > AGA[i]:
                                                cell.fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
                                            elif PH[i] <= PA[i] and AVGH[i] <= AVGA[i] and AGH[i] > AGA[i]:
                                                cell.fill = PatternFill(start_color="CCFF66", end_color="CCFF66", fill_type="solid")
                                            elif PH[i] <= PA[i] and AVGH[i] <= AVGA[i] and AGH[i] < AGA[i]:
                                                cell.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
                                            elif PH[i] >= PA[i] and AVGH[i] <= AVGA[i] and AGH[i] <= AGA[i]:
                                                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                                            elif PH[i] >= PA[i] and AVGH[i] <= AVGA[i] and AGH[i] > AGA[i]:
                                                cell.fill = PatternFill(start_color="B1A0C7", end_color="B1A0C7", fill_type="solid")
                    except:
                        pass                  

                    ########################################## LTN-1ST HOME-TEAM(A) AWAY-TEAM(A) ######################################
                    def LTN_1ST(home,away,H,A):
                        resH = []
                        count = 0
                        for row in (home):
                            for i in range (len(row)):
                                if row[i][0] == H and row[i][-1] == "L" and (row[i][3] == "0" or row[i][1] == "0") and count == 0:
                                    resH.append(True)
                                    count =+1
                                    break
                                elif row[i][0] != H:
                                    if i == 4:
                                        resH.append(False)
                                        break
                                    else:
                                        continue
                                else:
                                    resH.append(False)
                                    break
                        
                        resA = []
                        count = 0
                        for row in (away):
                            for i in range (len(row)):
                            
                                if row[i][0] == A and row[i][-1] == "L" and (row[i][3] == "0" or row[i][1] == "0") and count ==0:
                                    count += 1
                                    resA.append(True)
                                    break
                                elif row[i][0] != A:
                                    if i == 4:
                                        resA.append(False)
                                        break
                                    else:
                                        continue
                                else :
                                    resA.append(False)
                                    break
                        
                        Final_res = []
                        for i in range(len(resA)):
                            if resA[i] == True and resH[i] == True:
                                Final_res.append(True)
                            else:
                                Final_res.append(False)
                        return Final_res

                    Fres = LTN_1ST(home,away,'A','A')
                    Fres2 = LTN_1ST(home,away,'A','H')
                    Fres3 = LTN_1ST(home,away,'H','H')
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'Time' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if Fres[i] == True:
                                        cell.border = color1
                                    if Fres2[i] == True:
                                        cell.border = color2
                                    if Fres3[i] == True:
                                        cell.border = color2
                                    PROV_LTN1ST = True
                    except:
                        pass               

                    ############################################ CW ####################################################
                    def CW(home,away):
                        no = ["1-0","0-1","1-1","0-2","2-0","0-0"]
                        resH = []
                        for row in (home):
                            for i in range (len(row)):
                                if row[i][0] == 'A' and  row[i][1:4] not in no:
                                    resH.append(True)
                                    break
                                elif row[i][0] != 'A':
                                    if i == 4:
                                        resH.append(False)
                                        break
                                    else:
                                        continue
                                else:
                                    resH.append(False)
                                    break

                        resA = []
                        no = ["1-0","0-1","1-1","0-2","2-0","0-0"]
                        for row in (away):
                            for i in range (len(row)):
                                if row[i][0] == 'H' and  row[i][1:4] not in no:
                                    resA.append(True)
                                    break
                                elif row[i][0] != 'H':
                                    if i == 4:
                                        resA.append(False)
                                        break
                                    else:
                                        continue
                                else:
                                    resA.append(False)
                                    break

                        return  resH,resA

                    rH, rA = CW(home,away)
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=8, max_col=8):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'AVGH' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if rH[i] == True:
                                        cell.fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
                        i = -1
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=9, max_col=9):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'AVGA' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if rA[i] == True:
                                        cell.fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
                    except:
                        pass                
               
                    ########################################### DBT TASK ####################################                
                    def DBT(home,away,H,A):
                        yes = ['0-0','1-1']
                        resH = []
                        for row in (home):
                            for i in range (len(row)):
                                if row[i][0] == H and row[i][-1] == "D" and row[i][1:4] in yes:
                                    resH.append(True)
                                    break
                                elif row[i][0] != H:
                                    if i == 4:
                                        resH.append(False)
                                        break
                                    else:
                                        continue
                                else :
                                    resH.append(False)
                                    break

                        resA = []
                        for row in (away):
                            for i in range (len(row)):
                                if row[i][0] == A and row[i][-1] == "D" and row[i][1:4] in yes:
                                    resA.append(True)
                                    break
                                elif row[i][0] != A:
                                    if i == 4:
                                        resA.append(False)
                                        break
                                    else:
                                        continue
                                else:
                                    resA.append(False)
                                    break
                        Final_res = []
                        for i in range(len(resA)):
                            if resA[i] == True and resH[i] == True:
                                Final_res.append(True)
                            else:
                                Final_res.append(False)
                        return  Final_res

                    resDBT = DBT(home,away,"H",'H')
                    resDBT2 = DBT(home,away,"H",'A')
                    resDBT3 = DBT(home,away,"A",'H')
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=4):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'Home' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if resDBT[i] == True:
                                        cell.border = color4
                                    if resDBT2[i] == True:
                                        cell.border = color5
                                    if resDBT3[i] == True:
                                        cell.border = color6
                                    PROV_DBT = True
                    except:
                        pass

                    ################################################ BK TASK ####################################
                    def BK(home,away,H,A):
                        yes = ["1-0","2-0","3-0","4-0","0-1","0-2","0-3","0-4"]
                        resH = []
                        for row in (home):
                            for i in range (len(row)):
                                if row[i][0] == H and row[i][-1] == "W" and row[i][1:4] in yes:
                                    resH.append(True)
                                    break
                                elif row[i][0] != H:
                                    if i == 4:
                                        resH.append(False)
                                        break
                                    else:
                                        continue
                                else:
                                    resH.append(False)
                                    break

                        resA = []
                        for row in (away):
                            for i in range (len(row)):
                                if row[i][0] == A and row[i][-1] == "W" and row[i][1:4] in yes:
                                    resA.append(True)
                                    break
                                elif row[i][0] != A:
                                    if i == 4:
                                        resA.append(False)
                                        break
                                    else:
                                        continue
                                else:
                                    resA.append(False)
                                    break
                        Final_res = []
                        for i in range(len(resA)):
                            if resA[i] == True and resH[i] == True:
                                Final_res.append(True)
                            else:
                                Final_res.append(False)
                        return  Final_res

                    BKres = BK(home,away,"A",'A')
                    BKres2 = BK(home,away,"A",'H')
                    BKres3 = BK(home,away,"H",'H')
                    BKres4 = BK(home,away,"H",'A')
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=9, max_col=9):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'AVGA' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if BKres[i] == True:
                                        cell.border = color7
                                    if BKres2[i] == True:
                                        cell.border = color8
                    except:
                        pass

                    i = -1
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=8, max_col=8):
                        for cell in row:
                            if cell.value is not None and  cell.value != 'AVGH' and cell.value !='' and isinstance(cell.value,str) :
                                i += 1
                                if BKres[i] == True:
                                    cell.border = color9
                                if BKres2[i] == True:
                                    cell.border = color10
                    
                    ################################################ 3NR  NEED TO BE FIXED Prob is missing ####################################
                    def Three_NR(home,away):
                        resH = []
                        for row in home:
                            count = "" 
                            for i in range(len(row)):
                                if row[i][-1] == "L":
                                    count += row[i][-1]
                                else:
                                    count += row[i][-1]
                                
                            if "LLL" in count :
                                resH.append(True)
                            else:
                                resH.append(False)
                        
                        resA = []
                        for row in away:
                            count = ""
                            for i in range(len(row)):
                                if row[i][-1] == "L":
                                    count += row[i][-1]
                                else:
                                    count += row[i][-1]
                            if "LLL" in count :
                                resA.append(True)
                            else:
                                resA.append(False)

                        return resH,resA
                    TNRH, TNRA =Three_NR(home,away)
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6, max_col=6):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'GH' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if TNRH[i] == True and PH[i] > PA[i]:
                                        cell.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
                    except:
                        pass          

                    try:
                        i = -1
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=7, max_col=7):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'GA' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if TNRA[i] == True and PH[i] < PA[i]:
                                        cell.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
                    except:
                        pass  

                    ################################################ AF ####################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'Time' and cell.value !='' and isinstance(cell.value,str) :
                                    i += 1
                                    if float(AVGH[i]) + float(AVGA[i]) >= 4.8 and float(AVGH[i]) + float(AVGA[i]) <= 7.2:
                                        font = Font(color='FF0066',name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                                    if float(AVGH[i]) + float(AVGA[i]) >= 7.4 and float(AVGH[i]) + float(AVGA[i]) <= 12.0:
                                        font = Font(color='00FFFF',name='Arial Narrow', size=11, bold=True, italic=True)
                                        cell.font = font
                    except:
                        pass

                    def check_data(t):
                        home_score_check = ["2-1", "2-2", "3-1", "3-2", "3-3", "4-0", "4-1", "4-2", "4-3", "5-0", "5-1", "5-2", "5-3", "5-4"]
                        away_score_check = ["1-2", "2-2", "1-3", "2-3", "3-3", "0-4", "1-4", "2-4", "3-4", "0-5", "1-5", "2-5", "3-5", "4-5"]


                        home =[row[0:5] for row in t]
                        away =[row[5:] for row in t]
                        home_check = []
                        away_check = []
                        x = 0
                        for i in range(len(home)):
                            for elm in home[i]:
                                if elm[1:4] in home_score_check and elm[0] == "H" and elm[-1] != "L":
                                    x += 1
                                    #print(elm)
                                
                            if x >= 1:
                                home_check.append(True)
                                x = 0
                            else:
                                home_check.append(False)    
                        x = 0
                        for i in range(len(away)):
                            for elm in away[i]:
                                if elm[1:4] in away_score_check and elm[0] == "A" and elm[-1] != "L":
                                    #print("away",elm)
                                    x += 1
                                
                            if x >= 1:
                                away_check.append(True)
                                x = 0
                            else:
                                away_check.append(False)    
                        return home_check,away_check
                    homeC, awayA = check_data(raw_data)
                    ic(homeC)
                    ic(awayA)
                    inde = -1
                    
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
                            for cell in row:
                                try:
                                    if cell.value is not None and  cell.value != 'Date' and cell.value !='' and isinstance(cell.value,str) :
                                        inde += 1
                                        if homeC[inde] == True and awayA[inde]== False:
                                            cell.fill = PatternFill(start_color="FFCE33", end_color="FFCE33", fill_type="solid")
                                            if PH[inde] >= PA[inde]:
                                                if AVGH[inde] >= AVGA[inde]:
                                                    font = Font(color='9FFF05', bold=True ,italic=True)  # Red color in hexadecimal notation
                                                    cell.font = font
                                                else:
                                                    font = Font(color='E4080A', bold=True ,italic=True)  # Red color in hexadecimal notation
                                                    cell.font = font

                                        if homeC[inde] == False and awayA[inde]== True:
                                            cell.fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type="solid")
                                        if homeC[inde] == True and awayA[inde]== True:
                                            cell.fill = PatternFill(start_color="EFBCFE", end_color="EFBCFE", fill_type="solid")
                                except:
                                    pass
                    except:
                        pass

                    def check_dataa(t):
                        away_score_check = ["2-1", "2-2", "3-1", "3-2", "3-3", "4-0", "4-1", "4-2", "4-3", "5-0", "5-1", "5-2", "5-3", "5-4"]
                        home_score_check = ["1-2", "2-2", "1-3", "2-3", "3-3", "0-4", "1-4", "2-4", "3-4", "0-5", "1-5", "2-5", "3-5", "4-5"]

                        home =[row[0:5] for row in t]
                        away =[row[5:] for row in t]
                        home_check = []
                        away_check = []
                        x = 0
                        try:
                            for i in range(len(home)):
                                for elm in home[i]:
                                    if elm[1:4] in home_score_check and elm[0] == "A" and elm[-1] != "L":
                                        x += 1
                                        #print(elm)
                                    
                                if x >= 1:
                                    home_check.append(True)
                                    x = 0
                                else:
                                    home_check.append(False)
                        except:
                            pass    

                        x = 0
                        try:
                            for i in range(len(away)):
                                for elm in away[i]:
                                    if elm[1:4] in away_score_check and elm[0] == "H" and elm[-1] != "L":
                                        #print("away",elm)
                                        x += 1
                                    
                                if x >= 1:
                                    away_check.append(True)
                                    x = 0
                                else:
                                    away_check.append(False)
                        except:
                            pass 
                        return home_check,away_check
                    homeC, awayA = check_dataa(raw_data)
                    ic(homeC)
                    ic(awayA)
                    inde = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'League' and cell.value !='' and isinstance(cell.value,str) :
                                    inde += 1
                                    if homeC[inde] == True and awayA[inde]== False:
                                        cell.fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type="solid")
                                    if homeC[inde] == False and awayA[inde]== True:
                                        cell.fill = PatternFill(start_color="FFCE33", end_color="FFCE33", fill_type="solid")
                                    if homeC[inde] == True and awayA[inde]== True:
                                        cell.fill = PatternFill(start_color="EFBCFE", end_color="EFBCFE", fill_type="solid")
                    except:
                        pass

                    try:
                        for team_score_index in range(len(t1_avg)):    
                            new_avg.append([t1_avg[team_score_index],t2_avg[team_score_index]])
                        
                        thick_red_border = Border(left=Side(border_style='thick', color='FF0000'),
                            right=Side(border_style='thick', color='FF0000'),
                            top=Side(border_style='thick', color='FF0000'),
                            bottom=Side(border_style='thick', color='FF0000'))
                        po = 0
                        for cell in ws['D']:
                            cell.font = Font(name='Arial Narrow', size=11, bold=True, italic=True)
                        for cell in ws['B']:
                            cell.font = Font(name='Arial Narrow', size=11, bold=True, italic=True)
                    except:
                        pass
                    
                    if res != None:
                        po = 0
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=4):
                            for cell in row:
                                try:
                                    if cell.value is not None and  cell.value != 'Home' and cell.value !='' and isinstance(cell.value,str) :
                                        if res[po][0] == True and float(SDH[po]) <= 0.3099999:
                                            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                                            font = Font(color='FFFFFF',name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                        elif res[po][0] == True and float(SDH[po]) >= 0.31:
                                            cell.fill = PatternFill(start_color="9B9A9A", end_color="9B9A9A", fill_type="solid")
                                            font = Font(color='FFFFFF',name='Arial Narrow', size=11, bold=True, italic=True) 
                                            cell.font = font

                                        po += 1
                                except:
                                    pass
                        

                        po = 0
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
                            for cell in row:
                                try:
                                    if cell.value is not None and  cell.value != 'Away' and cell.value !='' and isinstance(cell.value,str) :
                                        if res[po][1] == True and float(SDA[po]) <= 0.3099999:
                                            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                                            font = Font(color='FFFFFF',name='Arial Narrow', size=11, bold=True, italic=True)
                                            cell.font = font
                                        elif res[po][1] == True and float(SDA[po]) >= 0.31:
                                            cell.fill = PatternFill(start_color="9B9A9A", end_color="9B9A9A", fill_type="solid")
                                            font = Font(color='FFFFFF',name='Arial Narrow', size=11, bold=True, italic=True) 
                                            cell.font = font

                                        po += 1
                                except:
                                    pass

                   

                    #SYM...
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=9, max_col=9):
                            for cell in row:
                                if cell.value is not None and  cell.value !='' and cell.value != 'AVGA' and isinstance(cell.value,str) :
                                    i += 1
                                    if (PROV_TIG and (float(GVA[i]) + float(GVH[i]) >= 1.74)):
                                        cell.fill = PatternFill(start_color="ABFFDD", end_color="ABFFDD", fill_type="solid")
                    except:
                        pass  
                    
                    Avgs =[]
                    Tavgs = []
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=8, max_col=8):
                        for cell in row:
                            try:
                                if cell.value is not None and cell.value !='' and cell.value != 'AVGA'and cell.value != 'AVGH':
                                    #print(cell.value)
                                    Avgs.append(cell.value)
                            except:
                                pass
                    
                    Avgs = [x for x in Avgs if not isinstance(x, float) or not math.isnan(x)]
                    if len(Avgs) % 2 != 0:
                        Avgs.append(0)
                    for team_score_index in range(0, len(Avgs), 2):
                        #print(Tavgs)
                        pair = [Avgs[team_score_index], Avgs[team_score_index + 1]]
                        Tavgs.append(pair)

                    Avgs = []
                    
                    k = 0
                    ind = 0                                    
                    Teams = []

                    column_letter = 'B'
                    for cell in ws[column_letter]:
                        cell.font = Font(name='Arial Nova', size=10, bold=False, italic=False)
                    for cell in ws['C']:
                        cell.font = Font(name='Arial Narrow', size=11, bold=True, italic=True)
                    for cell in ws['E']:
                        cell.font = Font(name='Arial Narrow', size=11, bold=True, italic=True)
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=25, max_col=25):
                        for cell in row:    
                            try:
                                if cell.value is not None and cell.value !='' and isinstance(cell.value, str) and  cell.value !="SUM" :
                                    Teams.append(cell.value)
                                    
                                    if float(Tavgs[ind][0]) + float(Tavgs[ind][1]) >= 7.4 and float(Tavgs[ind][0]) + float(Tavgs[ind][1]) >= 12.0:
                                        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") #Black SUM
                                        font = Font(color='4EEA10')  # Red color in hexadecimal notation
                                        cell.font = font
                                    ind += 1
                            except:
                                pass

                    Teams = []
                    ind = 0
                    
                    Teams = []
                    ind = 0 ###################### TO BE UPDATED 
                    thick_red_border = Border(left=Side(border_style='thick', color='C000F7'),
                          right=Side(border_style='thick', color='C000F7'),
                          top=Side(border_style='thick', color='C000F7'),
                          bottom=Side(border_style='thick', color='C000F7'))
                    x = -1
                      
                    Teams = []                   
                    ind = 0 
                    row_index = -1
                    thick_red_border = Border(left=Side(border_style='thick', color='E4080A'),
                          right=Side(border_style='thick', color='E4080A'),
                          top=Side(border_style='thick', color='E4080A'),
                          bottom=Side(border_style='thick', color='E4080A'))

                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
                        for cell in row:
                            try:
                                if cell.value is not None and  cell.value != 'Date'  and isinstance(cell.value,str) and '-' in cell.value:
                                    row_index += 1
                                    verified_team_1=False
                                    verified_team_2=False

                                    for team_score_index in range(1,5):
                                        L_and_0_T1 = 'L' ==  results[row_index][team_score_index][-1] and '0' in results[row_index][team_score_index][0:3]
                                        L_and_0_T2 = 'L' ==  results[row_index][team_score_index+5][-1] and '0' in results[row_index][team_score_index+5][0:3]

                                        if L_and_0_T1:
                                            verified_team_1=True
                                            
                                        if L_and_0_T2:
                                            verified_team_2=True

                                        if verified_team_1 and verified_team_2:    
                                            cell.border = thick_red_border
                                            break       # SWITCH TO BORDER
                            except:
                                pass
                     
                    def filtering_function_draw(results_team:list[list[str]])->list[list[bool]]:
                        boolean_list=[[False]*3 for _ in range(len(results_team))]
                        for row_index in range(len(results_team)):
                            verified_draw_t1=False
                            verified_draw_t2=False
                            verified_draw_3_or_4=False

                            for i in range(0,5):
                                team1_scores=results_team[row_index][i]
                                team2_scores=results_team[row_index][i+5]
                                
                                draw_same_order_t1=False
                                draw_same_order_t2=False

                                if(team1_scores[-1]=="D"):

                                    # condition for third function               
                                    draw_same_order_t1=True
                                    if(team1_scores[0]=="2" or team1_scores[0]=="1"):
                                        verified_draw_t1=True
                                    elif(team1_scores[0]=="3" or team1_scores[0]=="4"):
                                        verified_draw_3_or_4=True

                                    
                                if(team2_scores[-1]=="D"):
                                    draw_same_order_t2=True

                                    if(team2_scores[0]=="2" or team2_scores[0]=="1"):
                                        verified_draw_t2=True

                                    elif(team2_scores[0]=="3" or team2_scores[0]=="4"):
                                        verified_draw_3_or_4=True

                                # first function
                                if verified_draw_t1 and verified_draw_t2:
                                    boolean_list[row_index][0]=True
                                # second function
                                if verified_draw_3_or_4:
                                    boolean_list[row_index][1]=True
                                # third function
                                if draw_same_order_t1 and draw_same_order_t2:
                                    boolean_list[row_index][2]=True            
                            
                                
                        return boolean_list

                    final_res = filtering_function_draw(results)
                    p = -1
                    thick_red_border = Border(left=Side(border_style='thick', color='EB07FF'),
                          right=Side(border_style='thick', color='EB07FF'),
                          top=Side(border_style='thick', color='EB07FF'),
                          bottom=Side(border_style='thick', color='EB07FF')) 

                    # Apply cell coloring based on values in the 'W/L' column
                    try:
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=41):
                            for cell in row:
                                # Skip empty cells
                                if cell.value is not None:
                                    if cell.value == 'D':
                                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                                    elif cell.value == 'W':
                                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                                    elif cell.value == 'L':
                                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                    except:
                        pass
                                
                    c =  -1
                    z = 0
                    zz = 0
                    bn = 0
                    ############################################################## TTC ###################################################################################
                    TTC = []
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=23):   
                        for cell in row:
                            try:
                                if cell.value != '' and cell.value is not None and cell.value != 'CV Home' and cell.value != 'CV Away' and cell.value !='SD Home' and cell.value !='SD Away' and cell.value !='Goal Value A' and cell.value !='Goal Value H'and cell.value !='Goal Cost Home'and cell.value !='Goal Cost Away'and cell.value !='Prob.home'and cell.value !='Prob.Away' and isinstance(cell.value,str):
                                    c+=1
                                    
                                    if len(total_aux_list[z]) > 2 and cell.value[2:4] == total_aux_list[z][0] :
                                        if total_aux_list[z][1] == bn :
                                            font = Font(color='E3C901')  # Red color in hexadecimal notation
                                            cell.font = font
                                            cell.font = Font(color = cell.font.color.rgb,name='Arial Nova', size=11, bold=True, italic=True)
                                            
                                        elif total_aux_list[z][2] == bn:
                                            font = Font(color='E3C901')  # Red color in hexadecimal notation
                                            cell.font = font
                                            cell.font = Font(color = cell.font.color.rgb,name='Arial Nova', size=11, bold=True, italic=True)
                                        
                                        PROV_TTC = True

                                    bn+=1   
                                    if bn > 9:
                                        c= 0
                                        z+=1
                            except:
                                pass
                        bn = 0
                    ################################################################ SCR TTC preperation #############################################
                    c =  -1
                    z = 0
                    zz = 0
                    bn = 0
                    for i in range (len(total_aux_list)):
                        if len(total_aux_list[i])>=3:
                            TTC.append(True)
                        else:
                            TTC.append(False)
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=16, max_col=23):
                        for cell in row:
                            try:
                                if cell.value != '' and cell.value is not None and cell.value != 'CV Home' and cell.value != 'CV Away' and cell.value !='SD Home' and cell.value !='SD Away' and  cell.value !='Goal Cost Home'and cell.value !='Goal Cost Away'and cell.value !='Prob.home'and cell.value !='Prob.Away' and isinstance(cell.value,str):
                                    c+=1
                                    if len(total_aux_list[z]) > 2 and cell.value[2:4] == total_aux_list[z][0] :
                                        pass
                                    bn+=1
    
                                    if bn > 9:
                                        c= 0
                                        z+=1
                            except:
                                pass
                        bn = 0
                    ################################################################## SCR #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=24, max_col=24):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Early' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    if TTC[i] == True and (ZOYA[i] == True or ZOYA2[i]==True):
                                        font = Font(color='FF0000',bold=True, italic=True)  # Red color in hexadecimal notation
                                        cell.font = font
                                    PROV_SCR = True
                    except:
                        pass       
                    ################################################################## DP #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=26, max_col=26):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Live' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    if float(SDH[i]) + float(SDA[i]) >= 0 and float(SDH[i]) + float(SDA[i]) <= 0.27999999:
                                        cell.fill = PatternFill(start_color="AC75D5", end_color="AC75D5", fill_type="solid")  
                                        PROV_DP = True
                                    if float(SDH[i]) + float(SDA[i]) >= 0.28 and float(SDH[i]) + float(SDA[i]) <= 0.44999999:
                                        cell.fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")
                                        PROV_DP = True
                                    if float(SDH[i]) + float(SDA[i]) >= 0.45 and float(SDH[i]) + float(SDA[i]) <= 0.60999999:
                                        cell.fill = PatternFill(start_color="FFD347", end_color="FFD347", fill_type="solid")
                                        PROV_DP = True
                                    if float(SDH[i]) + float(SDA[i]) >= 0.61 and float(SDH[i]) + float(SDA[i]) <= 0.81999999:
                                        cell.fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")
                                        PROV_DP = True
                    except:
                        pass

                    ################################################################## 3DD and SCP #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=25, max_col=25):
                            for cell in row:
                                if cell.value is not None and  cell.value != 'SUM' and cell.value !='' and isinstance(cell.value,float) and len(str(cell.value)) >= 1 and str(cell.value) != "nan":
                                    i += 1
                                    ic(str(cell.value))
                                    if ECC[i] == True and (ZOYA[i] == True or ZOYA2[i] == True):
                                        font = Font(color='FF00FF', bold= True, italic=True)  # Red color in hexadecimal notation
                                        cell.font = font
                                    if PH[i] > PA[i]:
                                        x = float(GCH[i]) + float(CVH[i]) + 2894
                                        y = float(GCH[i]) + float(CVH[i]) + float(CVA[i])
                                        ic(x)
                                        ic(y)

                                        if float(GCA[i]) + float(CVH[i]) + float(CVA[i]) >=0 and float(GCA[i]) + float(CVH[i]) + float(CVA[i]) <=0.17999999999999:
                                            cell.fill = PatternFill(start_color="AC75D5", end_color="AC75D5", fill_type="solid")  
                                        if float(GCA[i]) + float(CVH[i]) + float(CVA[i]) >=0.18 and float(GCA[i]) + float(CVH[i]) + float(CVA[i]) <=0.2699999999999:
                                            cell.fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")  
                                        if float(GCA[i]) + float(CVH[i]) + float(CVA[i]) >=0.27 and float(GCA[i]) + float(CVH[i]) + float(CVA[i]) <=0.40999999:
                                            cell.fill = PatternFill(start_color="FFD347", end_color="FFD347", fill_type="solid")  
                                        if float(GCA[i]) + float(CVH[i]) + float(CVA[i]) >=0.41 and float(GCA[i]) + float(CVH[i]) + float(CVA[i]) <=0.91999999:
                                            cell.fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")  
                                    elif PH[i] <= PA[i]:
                                        if float(GCH[i]) + float(CVH[i]) + float(CVA[i]) >=0 and float(GCH[i]) + float(CVH[i]) + float(CVA[i]) <=0.179999999999999:
                                            cell.fill = PatternFill(start_color="AC75D5", end_color="AC75D5", fill_type="solid")  
                                        if float(GCH[i]) + float(CVH[i]) + float(CVA[i]) >=0.18 and float(GCH[i]) + float(CVH[i]) + float(CVA[i]) <=0.2699999999999:
                                            cell.fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")  
                                        if float(GCH[i]) + float(CVH[i]) + float(CVA[i]) >=0.27 and float(GCH[i]) + float(CVH[i]) + float(CVA[i]) <=0.40999999:
                                            cell.fill = PatternFill(start_color="FFD347", end_color="FFD347", fill_type="solid") 
                                        if float(GCH[i]) + float(CVH[i]) + float(CVA[i]) >=0.41 and float(GCH[i]) + float(CVH[i]) + float(CVA[i]) <=0.91999999:
                                            cell.fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")  
                                    PROV_3DD = True
                    except:
                        pass

                    thick_Black_border = Border(left=Side(border_style='thick', color='000000'),
                          right=Side(border_style='thick', color='000000'),
                          top=Side(border_style='thick', color='000000'),
                          bottom=Side(border_style='thick', color='000000'))
                    thick_Pink_border = Border(left=Side(border_style='thick', color='FF0066'),
                          right=Side(border_style='thick', color='FF0066'),
                          top=Side(border_style='thick', color='FF0066'),
                          bottom=Side(border_style='thick', color='FF0066'))
                    thick_Blue_border = Border(left=Side(border_style='thick', color='0000FF'),
                          right=Side(border_style='thick', color='0000FF'),
                          top=Side(border_style='thick', color='0000FF'),
                          bottom=Side(border_style='thick', color='0000FF'))
                    thick_Gold_border = Border(left=Side(border_style='thick', color='FACA00'),
                          right=Side(border_style='thick', color='FACA00'),
                          top=Side(border_style='thick', color='FACA00'),
                          bottom=Side(border_style='thick', color='FACA00'))
                    ################################################################## F-09 #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Away' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    if SDH[i] > SDA[i] :
                                        if   float(SDA[i]) * 1.25 < float(SDH[i]) and float(SDA[i]) * 2.17 > float(SDH[i]) and float(SDA[i]) <= 0.31 and float(SDA[i]) >= 0.11 :
                                            if (ZOYA[i] == True or ZOYA2[i] == True):
                                                font = Font(color='FFFF0000', name='Arial Narrow', size=11, bold=True, italic=True)
                                                cell.font = font
                                                #cell.border = thick_Black_border
                                            elif (Pink_Zoya[i] == True or Pink_Zoya2[i] == True):
                                                font = Font(color='FFFF0000', name='Arial Narrow', size=11, bold=True, italic=True)
                                                cell.font = font
                                                #cell.border = thick_Pink_border
                                            PROV_F09 = True
                                        
                                    else:  
                                        if   float(SDH[i]) * 1.25 < float(SDA[i]) and float(SDH[i]) * 2.17 > float(SDA[i]) and float(SDH[i]) <= 0.31 and float(SDH[i]) >= 0.11:  # ZOYAF_09[i] == True and
                                            if (ZOYA[i] == True or ZOYA2[i] == True) :
                                                font = Font(color='FFFF0000', name='Arial Narrow', size=11, bold=True, italic=True)
                                                cell.font = font
                                                #cell.border = thick_Black_border
                                            elif (Pink_Zoya[i] == True or Pink_Zoya2[i] == True):
                                                font = Font(color='FFFF0000', name='Arial Narrow', size=11, bold=True, italic=True)
                                                cell.font = font
                                                #cell.border = thick_Pink_border
                                            PROV_F09 = True
                    except:
                        pass                
                    ################################################################## PAV #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = str(float(GCH[i]) *1.04)
                                    val2 = str(float(GCH[i]) *0.96)
                                    if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_PAV = True
                    except:
                        pass                   

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = str(float(GCA[i]) *1.04)
                                    val2 = str(float(GCA[i]) *0.96)
                                    if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_PAV = True
                    except:
                        pass
                        
                    ################################################################## HQO #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *1.04)
                                    val2 = str(float(GCH[i]) *0.96)
                                    if  val2[val2.index('.')+1:val2.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="FFAFFF", end_color="FFAFFF", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_HQO = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = str(float(GCA[i]) *1.04)
                                    val2 = str(float(GCA[i]) *0.96)
                                    if  val2[val2.index('.')+1:val2.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="FFAFFF", end_color="FFAFFF", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_HQO = True
                    except:
                        pass 
                    ################################################################## CIS #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *2.4)
                                    val2 = str(float(GCH[i]) *2.51)
                                    if  val[val.index('.')+1:val.index('.')+3] <= HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3] <= val2[val2.index('.')+1:val2.index('.')+3]:
                                        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_CIS = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCA[i]) *2.4)
                                    val2 = str(float(GCA[i]) *2.51)
                                    if  val[val.index('.')+1:val.index('.')+3] <= AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] <= val2[val2.index('.')+1:val2.index('.')+3]:
                                        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_CIS = True
                    except:
                        pass 

                    ################################################################## ASY #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != "Goal Cost Home" and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(HGD[i]) *2.4)
                                    val2 = str(float(HGD[i]) *2.51)

                                    if  val[val.index('.')+1:val.index('.')+3] <= GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3] <= val2[val2.index('.')+1:val2.index('.')+3]:
                                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_ASY = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != "Goal Cost Away" and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(AGD[i]) *2.4)
                                    val2 = str(float(AGD[i]) *2.51)

                                    if  val[val.index('.')+1:val.index('.')+3] <= GCA[i][GCA[i].index('.')+1:GCA[i].index('.')+3] <= val2[val2.index('.')+1:val2.index('.')+3]:
                                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_ASY = True
                    except:
                        pass 

                    ################################################################## CEO #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *2.80)
                                    #val2 = str(float(GCH[i]) *0.96)
                                    if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="FF0066", end_color="FF0066", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_CEO = True
                    except:
                        pass              

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCA[i]) *2.80)
                                    if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="FF0066", end_color="FF0066", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_CEO = True
                    except:
                        pass
                    
                    ################################################################## RKT #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *3.5999)
                                    if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="00FF99", end_color="00FF99", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_CEO = True
                    except:
                        pass              

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCA[i]) *3.5999)
                                    if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="00FF99", end_color="00FF99", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_CEO = True
                    except:
                        pass

                    ################################################################## AAO #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(HGD[i]) *2.80)
                                    if val[val.index('.')+1:val.index('.')+3] == GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="5B114D", end_color="5B114D", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_AAO = True
                    except:
                        pass  
                                 
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(AGD[i]) *2.80)
                                    if val[val.index('.')+1:val.index('.')+3] == GCA[i][GCA[i].index('.')+1:GCA[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="5B114D", end_color="5B114D", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_AAO = True
                    except:
                        pass
                    
                    ################################################################## NRG #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(HGD[i]) *0.6945)
                                    if val[val.index('.')+1:val.index('.')+3] == GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_AAO = True
                    except:
                        pass  
                                 
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(AGD[i]) *0.6945)
                                    if val[val.index('.')+1:val.index('.')+3] == GCA[i][GCA[i].index('.')+1:GCA[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_AAO = True
                    except:
                        pass

                    ################################################################## THK #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(HGD[i]) *1.3889)
                                    if val[val.index('.')+1:val.index('.')+3] == GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="6600FF", end_color="6600FF", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_AAO = True
                    except:
                        pass  
                                 
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(AGD[i]) *1.3889)
                                    if val[val.index('.')+1:val.index('.')+3] == GCA[i][GCA[i].index('.')+1:GCA[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="6600FF", end_color="6600FF", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_AAO = True
                    except:
                        pass

                    ################################################################## COO #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *2.88)
                                    if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="7DFF7D", end_color="7DFF7D", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_COO = True
                    except:
                        pass
                                    
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCA[i]) *2.88)
                                    if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="7DFF7D", end_color="7DFF7D", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_COO = True
                    except:
                        pass 
                        
                    ################################################################## THU #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *3)
                                    if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_THU = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCA[i]) *3)
                                    if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_THU = True
                    except:
                        pass     

                    ################################################################## CIN #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != "Goal Value A" and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val2 = str(float(GCH[i]) *0.96)
                                    if  val2[val2.index('.')+1:val2.index('.')+3] == GVA[i][GVA[i].index('.')+1:GVA[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="FF99CC", end_color="FF99CC", fill_type="solid")  
                                        #SCA_Check = True
                                        PROV_CIN = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != "Goal Value H" and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val2 = str(float(GCA[i]) *0.96)
                                    if  val2[val2.index('.')+1:val2.index('.')+3] == GVH[i][GVH[i].index('.')+1:GVH[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="FF99CC", end_color="FF99CC", fill_type="solid")  
                                        #SCA_Check = True
                                        PROV_CIN = True
                    except:
                        pass 

                     ################################################################## HSS #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value A' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *1.04)
                                    val2 = str(float(GCH[i]) *0.96)
                                    if val[val.index('.')+1:val.index('.')+3] == GVA[i][GVA[i].index('.')+1:GVA[i].index('.')+3]:# or val2[val2.index('.')+1:val2.index('.')+3] == GVA[i][GVA[i].index('.')+1:GVA[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")  
                                        #SCA_Check = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value H' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCA[i]) *1.04)
                                    val2 = str(float(GCA[i]) *0.96)
                                    if val[val.index('.')+1:val.index('.')+3] == GVH[i][GVH[i].index('.')+1:GVH[i].index('.')+3]:# or val2[val2.index('.')+1:val2.index('.')+3] == GVH[i][GVH[i].index('.')+1:GVH[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")  
                                        #SCA_Check = True
                                        PROV_HSS = True
                    except:
                        pass 

                    ################################################################## HOY #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != "Goal Value A" and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *1.12)
                                    if val[val.index('.')+1:val.index('.')+3] == GVA[i][GVA[i].index('.')+1:GVA[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")  
                                        #SCA_Check = True
                                        PROV_HOY = True
                    except:
                        pass            

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != "Goal Value H" and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCA[i]) *1.12)
                                    if val[val.index('.')+1:val.index('.')+3] == GVH[i][GVH[i].index('.')+1:GVH[i].index('.')+3] :
                                        cell.fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_HOY = True
                    except:
                        pass
                    
                    ################################################################## WHP #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != "Goal Value A" and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) * 0.88)
                                    
                                    if val[val.index('.')+1:val.index('.')+3] == GVA[i][GVA[i].index('.')+1:GVA[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")
                                        SCA_Check = True
                                        PROV_WHP = True
                    except:
                        pass                 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != "Goal Value H" and cell.value != "" and isinstance(cell.value,str):
                                    i += 1          
                                    val = str(float(GCA[i]) * 0.88)
                                    if val[val.index('.')+1:val.index('.')+3] == GVH[i][GVH[i].index('.')+1:GVH[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")
                                        #SCA_Check = True
                                        PROV_WHP = True
                    except:
                        pass 

                    ################################################################## VLN #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCH[i]) *1.12)
                                    val2 = str(float(GCH[i]) *0.88)
                                    if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  
                                        #SCA_Check = True
                                    if val2[val2.index('.')+1:val2.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="47ABCC", end_color="47ABCC", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_VLN = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str(float(GCA[i]) *1.12)
                                    val2 = str(float(GCA[i]) *0.88)
                                    if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  
                                        #SCA_Check = True
                                        PROV_VLN = True
                                    if val2[val2.index('.')+1:val2.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="47ABCC", end_color="47ABCC", fill_type="solid") 
                                        SCA_Check = True
                                        PROV_VLN = True
                    except:
                        pass 


                    ################################################################## HOC #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCH[i]) *2 )*1.04)                                
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3]:
                                            cell.fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")  
                                            #SCA_Check = True
                                            PROV_HOC = True
                    except:
                        pass                     

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCA[i]) *2 )*1.04)
                        
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] :
                                            cell.fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")
                                            #SCA_Check = True
                                            PROV_HOC = True
                    except:
                        pass 

                    ################################################################## HSY #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCH[i]) *2 )*0.96)
                                    
                                    if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="588AEE", end_color="588AEE", fill_type="solid")
                                        #SCA_Check = True
                                        PROV_HSY = True
                    except:
                        pass                 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCA[i]) *2 )*0.94)
                                    
                                    if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3]:
                                        cell.fill = PatternFill(start_color="588AEE", end_color="588AEE", fill_type="solid")
                                        #SCA_Check = True
                                        PROV_HSY = True
                    except:
                        pass 

                    ################################################################## OFF #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCH[i]) *2 )*1.12)
                                    
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3]:
                                            cell.fill = PatternFill(start_color="A50021", end_color="A50021", fill_type="solid")  
                                            #SCA_Check = True
                    except:
                        pass                     

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCA[i]) *2 )*1.12)
                        
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] :
                                            cell.fill = PatternFill(start_color="A50021", end_color="A50021", fill_type="solid")
                                            #SCA_Check = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCH[i]) *2 )*0.88)
                                    
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == HGD[i][HGD[i].index('.')+1:HGD[i].index('.')+3]:
                                            cell.fill = PatternFill(start_color="C8A200", end_color="C8A200", fill_type="solid")  
                                            #SCA_Check = True
                    except:
                        pass                    

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCA[i]) *2 )*0.88)
                        
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == AGD[i][AGD[i].index('.')+1:AGD[i].index('.')+3] :
                                            cell.fill = PatternFill(start_color="C8A200", end_color="C8A200", fill_type="solid")
                                            #SCA_Check = True
                    except:
                        pass 


                    ################################################################## BNG #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'HGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCH[i]) *2 )*1.04)                                    
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GVA[i][GVA[i].index('.')+1:GVA[i].index('.')+3]:
                                            cell.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")  
                                            SCA_Check = True
                                            PROV_BNG = True
                    except:
                        pass 
                                        

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(GCA[i]) *2 )*1.04)                        
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3] :
                                            cell.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
                                            SCA_Check = True
                                            PROV_BNG = True
                    except:
                        pass 
                
                    ################################################################## GD's-Z #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'AGD' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    if float(AGD[i]) == 0 and (float(GVA[i]) + float (GVH[i])) >= 2.6:
                                        cell.border = thick_red_border
                                        SCA_Check = True
                                        PROV_GDSZ = True
                    except:
                        pass 

                    ################################################################## TWI's #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value H' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = str(float(HGD[i]))
                                    val2 = str(float(GVA[i]))
                                    if val[val.index('.')+1:val.index('.')+3]  == val2[val2.index('.')+1:val2.index('.')+3]:
                                        cell.fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
                                        SCA_Check = True  
                                        PROV_TWIS = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value H' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = str(float(AGD[i]))
                                    val2 = str(float(GVH[i]))
                                    if val[val.index('.')+1:val.index('.')+3]  == val2[val2.index('.')+1:val2.index('.')+3]:
                                        cell.fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")  
                                        SCA_Check = True
                                        PROV_TWIS = True
                    except:
                        pass 

                    ################################################################## PTN #############################################################
                    color111 = Border(left=Side(border_style='thick', color='0000FF'),
                          right=Side(border_style='thick', color='0000FF'),
                          top=Side(border_style='thick', color='0000FF'),
                          bottom=Side(border_style='thick', color='0000FF'))
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value A' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = str(float(HGD[i]))
                                    val2 = str(float(AGD[i]))
                                    if float(val) != 0 and float(val2) != 0:
                                        if val[val.index('.')+1:val.index('.')+3]  == val2[val2.index('.')+1:val2.index('.')+3]:
                                            cell.border = color111
                                            SCA_Check = True
                                            PROV_PTN = True
                    except:
                        pass               
                    ################################################################## RV #############################################################

                    #i = -1
                    #print('################################## RV #########################')
                    #try:
                    #    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                    #        for cell in row:
                    #            if cell.value != '' and cell.value is not None and cell.value != 'Goal Value H' and cell.value != "" and isinstance(cell.value,str):
                    #                i += 1
                    #                check = [0,1.0,-1.0]
                    #                if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check :
                    #                    val = str(float(HGD[i]))
                    #                    #print(GVA[i][GVA[i].index('.')+1:])
                    #                    #print(val[val.index('.')+3:val.index('.')+10])
                    #                    
                    #                    
                    #                    if val[val.index('.')+3:val.index('.')+10]  in GVA[i][GVA[i].index('.')+3:]:
                    #                        cell.fill = PatternFill(start_color="33CCCC", end_color="33CCCC", fill_type="solid") 
                    #                        SCA_Check = True 
                    #except:
                    #    pass 

                    #i = -1
                    #try:
                    #    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                    #        for cell in row:
                    #            if cell.value != '' and cell.value is not None and cell.value != 'Goal Value A' and cell.value != "" and isinstance(cell.value,str):
                    #               i += 1
                    #                check = [0,1.0,-1.0]
                    #                if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check :
                    #                    val = str(float(AGD[i]))
                    #                    
                    #                    if val[val.index('.')+3:val.index('.')+10]  in GVH[i][GVH[i].index('.')+3:]:
                    #                        cell.fill = PatternFill(start_color="33CCCC", end_color="33CCCC", fill_type="solid") 
                    #                        SCA_Check = True
                    #except:
                    #    pass 


                    ################################################################## AH #############################################################

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=22, max_col=22):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'SD Home' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    check = [0,1.0,-1.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check :
                                        val = str(float(HGD[i]))
                                        
                                        if val[val.index('.')+3:val.index('.')+10]  in str(PH[i])[str(PH[i]).index('.')+3:]:
                                            cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  
                                            SCA_Check = True
                                            PROV_AH = True
                                        if AGD[i][AGD[i].index('.')+3:AGD[i].index('.')+10]  in PA[i][PA[i].index('.')+3:]:
                                            cell.fill = PatternFill(start_color="85DFFF", end_color="85DFFF", fill_type="solid")  
                                            SCA_Check = True
                                            PROV_AH = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=22, max_col=22):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'SD Home' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    check = [0,1.0,-1.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check :
                                        val = str(float(AGD[i]))
                                        
                                        if val[val.index('.')+3:val.index('.')+10]  in PA[i][PA[i].index('.')+3:]:
                                            cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") 
                                            SCA_Check = True
                                            PROV_AH = True
                                        if HGD[i][HGD[i].index('.')+3:HGD[i].index('.')+10]  in PH[i][PH[i].index('.')+3:]:
                                            cell.fill = PatternFill(start_color="85DFFF", end_color="85DFFF", fill_type="solid") 
                                            SCA_Check = True
                                            PROV_AH = True
                    except:
                        pass 

                    ################################################################## SK #############################################################

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=21, max_col=21):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Prob.Away' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = PH[i]
                                    if str(PH[i][PH[i].index('.')+1:PH[i].index('.')+10])  in str(GVA[i][GVA[i].index('.')+1:]):
                                        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  
                                        font = Font(color='FFFFFF')  # Red color in hexadecimal notation
                                        cell.font = font
                                        if (float(PH[i])<=0.2 or float(PA[i])<=0.2):
                                            cell.fill = PatternFill(start_color="85DFFF", end_color="85DFFF", fill_type="solid")
                                            SCA_Check = True
                                            PROV_SK = True
                    except:
                        pass             

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=21, max_col=21):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Prob.Away' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = PA[i]
                                    if str(PA[i][PA[i].index('.')+1:PA[i].index('.')+10])  in str(GVH[i][GVH[i].index('.')+1:]):
                                        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") 
                                        font = Font(color='FFFFFF')  # Red color in hexadecimal notation
                                        cell.font = font
                                        if (float(PH[i])<=0.2 or float(PA[i])<=0.2):
                                            cell.fill = PatternFill(start_color="85DFFF", end_color="85DFFF", fill_type="solid")
                                            SCA_Check = True
                                            PROV_SK = True
                    except:
                        pass 
                    ################################################################## NX #############################################################

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=23, max_col=23):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'SD Away' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = PH[i]
                                    check = [0,1.0,-1.0]
                                    
                                    if str(PH[i][PH[i].index('.')+1:PH[i].index('.')+10])  in str(HGD[i][HGD[i].index('.')+1:])and (float(PH[i])<=0.2 or float(PA[i])<=0.2)and float(AGD[i]) - float(HGD[i]) not in check:
                                        cell.fill = PatternFill(start_color="403151", end_color="403151", fill_type="solid")  
                                        font = Font(color='FFFFFF')  # Red color in hexadecimal notation
                                        cell.font = font
                                        SCA_Check = True
                                        PROV_NX = True
                    except:
                        pass             

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=23, max_col=23):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'SD Away' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    val = PA[i]
                                    check = [0,1.0,-1.0]
                                    if str(PA[i][PA[i].index('.')+1:PA[i].index('.')+10])  in str(AGD[i][AGD[i].index('.')+1:])and (float(PH[i])<=0.2 or float(PA[i])<=0.2)and float(AGD[i]) - float(HGD[i]) not in check:
                                        cell.fill = PatternFill(start_color="403151", end_color="403151", fill_type="solid") 
                                        font = Font(color='FFFFFF')  # Red color in hexadecimal notation
                                        cell.font = font
                                        SCA_Check = True
                                        PROV_NX = True
                    except:
                        pass 

                    ################################################################## HCO #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value H' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(HGD[i]) *2 )*1.04)
                                    
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3]:
                                            cell.fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")  
                                            SCA_Check = True
                                            PROV_HCO = True
                    except:
                        pass                     

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value H' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(HGD[i]) *2 )*0.96)
                        
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3] :
                                            cell.fill = PatternFill(start_color="FFD901", end_color="FFD901", fill_type="solid")
                                            SCA_Check = True
                                            PROV_HCO = True
                    except:
                        pass 


                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value A' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(AGD[i]) *2 )*1.04)
                                    
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCA[i][GCA[i].index('.')+1:GCA[i].index('.')+3]:
                                            cell.fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")  
                                            PROV_HCO = True
                    except:
                        pass                     

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value A' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(AGD[i]) *2 )*0.96)
                        
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCA[i][GCA[i].index('.')+1:GCA[i].index('.')+3] :
                                            cell.fill = PatternFill(start_color="FFD901", end_color="FFD901", fill_type="solid")
                                            PROV_HCO = True
                    except:
                        pass                     


                    ################################################################## BK09 and QM #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value H' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(HGD[i]) *2 )*1.12)
                                    
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3]:
                                            cell.fill = PatternFill(start_color="A50021", end_color="A50021", fill_type="solid")  
                                            SCA_Check = True
                                            PROV_BK09 = True
                    except:
                        pass                         

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=15):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value H' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(HGD[i]) *2 )*0.88)                        
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCH[i][GCH[i].index('.')+1:GCH[i].index('.')+3] :
                                            cell.fill = PatternFill(start_color="33CC33", end_color="33CC33", fill_type="solid")
                                            SCA_Check = True
                                            PROV_BK09 = True
                    except:
                        pass 

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value A' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(AGD[i]) *2 )*1.12)
                                    
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCA[i][GCA[i].index('.')+1:GCA[i].index('.')+3]:
                                            cell.fill = PatternFill(start_color="A50021", end_color="A50021", fill_type="solid")  
                                            PROV_QM = True
                    except:
                        pass                     

                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Goal Value A' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    
                                    val = str((float(AGD[i]) *2 )*0.88)
                        
                                    check = [0,1.0,-1.0,2.0,-2.0]
                                    if float(AGD[i]) - float(HGD[i]) not in check or float(HGD[i]) - float(AGD[i]) not in check:
                                        if val[val.index('.')+1:val.index('.')+3] == GCA[i][GCA[i].index('.')+1:GCA[i].index('.')+3] :
                                            cell.fill = PatternFill(start_color="33CC33", end_color="33CC33", fill_type="solid")
                                            PROV_QM = True
                    except:
                        pass 

                    ################################################################## SCA #############################################################
                    i = -1
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=26, max_col=26):
                            for cell in row:
                                if cell.value != '' and cell.value is not None and cell.value != 'Live' and cell.value != "" and isinstance(cell.value,str):
                                    i += 1
                                    if (ZOYA[i] == True or ZOYA2[i] == True) or (Pink_Zoya[i] == True or Pink_Zoya2[i] == True) and SCA_Check == True :
                                        font = Font(color='0000FF',bold=True,italic=True)  # Red color in hexadecimal notation
                                        cell.font = font
                                        PROV_SCA = True
                    except:
                        pass             

                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=39, max_col=39):
                        for cell in row:
                            try:
                                if cell.value != '' and cell.value is not None and cell.value != 'Team' and cell.value != "" and isinstance(cell.value,str):
                                    if '(H)' in cell.value:
                                        font = Font(color='FE9900')  # Red color in hexadecimal notation
                                        cell.font = font
                                    elif '(A)' in cell.value:
                                        font = Font(color='02E1E9')  # Red color in hexadecimal notation
                                        cell.font = font
                            except:
                                pass

                    # Auto-adjust column widthss
                    for column in ws.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        try:
                            max_length = max(len(str(cell.value)) for cell in column)
                            adjusted_width = (max_length + 2)
                            ws.column_dimensions[column[0].column_letter].width = adjusted_width
                        except:
                            pass

                    try:
                        ic(TTC)
                        ic(ZOYA)
                    except:
                        pass
                    # Center align all cells
                    try:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                    except:
                        pass

                         # Save the workbook to an Excel file
                    date = act_data[0]['Date'].replace('-', '_')  # Extracting the first date from the list
                    file_path = f'{date}.xlsx'

                    # Updates By Wassim Karaouli (rearraging columns order)
                    new_sheet_order = ['Time', 'Date', 'League', 'Home vs Guest', 'Goal Value A', 'Goal Value H', 'Goal Cost Home', 'Goal Cost Away', 'Score(Finished)', 'SD', 'CV', 'Over', 'STAKE POOL', 'Early', 'SUM', 'Live', 'goal1', 'goal2', 'goal3', 'goal4', 'goal5', 'goal6', 'goal7', 'goal8', 'goal9', 'goal10', 'Prob.home', 'Prob.Away', 'Team', 'Score', 'W/L', 'AVG']
                
                    # Check if the file already exists
                    if os.path.exists(file_path):
                        print(f'Error: File {file_path} already exists.')
                        input('Please delete or move such file and press enter to retry')
                    else:
                        wb.save(file_path)
                        print(f"Excel file '{file_path}' created successfully.")
                    
                    time.sleep(2)            
                    # Read Excel file into a pandas DataFrame
                    #new_order = ['SD', 'CV','Time', 'Date', 'League', 'Home vs Guest', 'Goal Value A', 'Goal Value H', 'Goal Cost Home', 'Goal Cost Away', 'Score(Finished)' ,'Over', 'STAKE POOL', 'Early', 'SUM', 'Live', 'goal1', 'goal2', 'goal3', 'goal4', 'goal5', 'goal6', 'goal7', 'goal8', 'goal9', 'goal10', 'Prob.home', 'Prob.Away', 'Team', 'Score', 'W/L', 'AVG']
                    excel_file = file_path
                    df = pd.read_excel(excel_file)
                    #df = df[new_order]
                    #print(df)

                    # Identify rows with missing data in the first column
                    rows_with_missing_data = df[df.iloc[:, 0].isnull()]

                    # Get the row indices with missing data
                    indices_with_missing_data = rows_with_missing_data.index.tolist()

                    # Load the Excel file using openpyxl for further manipulation
                    wb = load_workbook(excel_file)
                    header_row = df.columns.tolist()
                    #print("bad_games -------->",bad_games)
                    
                    # Iterate through the identified rows and set the outline level to make them collapsible
                    for idx in indices_with_missing_data:
                        ws = wb.active  # Assuming you are working with the active sheet
                        row_number = idx + 2  # Excel rows are 1-indexed, DataFrame rows are 0-indexed
                        ws.row_dimensions[row_number].outline_level = 2  # Set the outline level to 2 or higher
                    
                    try:
                        os.remove(file_path)
                        #print(f'The file "{file_path}" has been successfully removed.')
                    except FileNotFoundError:
                        print(f'The file "{file_path}" does not exist.')
                    except Exception as e:
                        print(f'An error occurred: {e}')
                    # Save the modified Excel file
                    wb.save(file_path)
                    
            except Exception as e:
                print(f'Creating Excel File: {e}')
                pass
            driver.quit()
            #sys.exit()
            #print("bad_games -------->",bad_games)
            
        except Exception as e:
            message = 'an Error occured with Farhan Pirzada\'s script message him on upwork or email F.pirzada7@gmail.com\n'
            
    def dummy_automate(self):
        r = random.random()
        if r < 0.2:
            # Wait a random amount of time between 1 and 10 seconds and throw an error.
            time.sleep(random.uniform(1, 10))
            raise Exception(
                random.choice([
                    'I dont feel like doing any work today.',
                    'Ka-pow!',
                    'Boom!',
                    'Have you tried turning it off and on again?',
                    'Try with sudo. It may or may not work.',
                ])
            )
        else:  # 0.2 <= r < 1.0
            # "Correct" operation.
            while True:
                time.sleep(1)
                if self.requested_stop:
                    sys.exit()
lock = threading.Lock()

class Simulation:
    def __init__(self, url, run_time, n_threads, thread_start_delay, proxies_file, chunks):
        self.url = url
        self.run_time = run_time
        self.n_threads = n_threads
        self.thread_start_delay = thread_start_delay
        # self.proxies = self.load_proxies(proxies_file)
        self.proxies = proxies_file
        # self.combo = self.load_combos(combo_file)
        self.combo = chunks
        self.threads = []

    @staticmethod
    def load_combos(combo_file):
        combos = []
        combo_file = open(combo_file, "r")
        for line in combo_file.readlines():
            combos.append(line)
        return combos

    @staticmethod
    def load_proxies(proxies_file):
        proxies_file = open(proxies_file, "r")
        proxies = []
        for line in proxies_file.readlines():
            ip, port, user, password = [part.strip() for part in line.split(':')]
            proxy_url = f'{user}:{password}@{ip}:{port}'
            proxies.append(proxy_url)
        return proxies

    @staticmethod
    def print_runtime_stats(threads):
        running = sum(1 for t in threads if t.is_alive())
        failed = sum(1 for t in threads if t.failed)
        if running == 0:
            raise SystemExit

    def start(self):
        # Create threads but don't start them yet.
        self.threads = [
            SimulationThread(
                simulation=self,
                thread_id=thread_id,
                url=self.url,
                # proxy=self.proxies[thread_id],
                # combo=self.combo[thread_id],
                proxy = self.proxies,
                combo = self.combo,
                run_time=self.run_time
            )
            for thread_id in range(self.n_threads)
        ]
        
        # Set up our timer.
        start_time = time.time()
        finish_time = start_time + (self.run_time * 60 * 60)
        # Start threads.
        for thread in self.threads:
            thread.start()
            time.sleep(self.thread_start_delay)

        # Wait for the simulation run time to elapse.
        logging_timer = 0
        while time.time() < finish_time:
            time.sleep(1)
            
            logging_timer += 130
            if logging_timer % STAT_LOGGING_INTERVAL == 0:
                self.print_runtime_stats(self.threads)        

    def restart_thread(self, old_thread):
        logging.info(f'Restarting thread {thread.thread_id}')
        
        old_thread_index = self.threads.find(thread)
        new_thread = SimulationThread(
            simulation=self,
            thread_id=thread.thread_id,
            url=old_thread.url,
            proxy=old_thread.proxy,
            combo=old_thread.combo,
            run_time=old_thread.run_time,
        )
        self.threads[old_thread_index] = new_thread
        new_thread.start()

def extract_day_names(datesdict):
    day_names = [day_tuple[0] for day_tuple in datesdict.values()]
    return day_names


def puxa_datas():
    league = False
    datesdict = {}
    cookies = { "Time_Zone": "10" }
    soup = bs(requests.get('https://www.goaloo18.com/football/fixture', cookies=cookies).text, 'html.parser')
    dates = soup.select('ul[class="timeBox"] li')
    #dates = date_list.select('li')
    last_date = 0
    counter = 0
    for link in dates:
        date = link.get_text().strip().split()[0]
        date = f"{date[:3]} {date[3:]}"
        href = link.attrs['onclick'].split('"')[1]
        counter += 1
        datesdict[counter] = (date, href)
    return(extract_day_names(datesdict))


def puxa_dias_com_links():
    league = False
    datesdict = {}
    cookies = { "Time_Zone": "10" }
    soup = bs(requests.get('https://www.goaloo18.com/football/fixture', cookies=cookies).text, 'html.parser')
    dates = soup.select('ul[class="timeBox"] li')
    last_date = 0
    counter = 0
    for link in dates:
        date = link.get_text().strip().split()[0]
        date = f"{date[:3]} {date[3:]}"
        href = link.attrs['onclick'].split('"')[1]
        counter += 1
        datesdict[counter] = (date, href)
    return(datesdict)


def get_link_by_day(day_name, datesdict):
    for key, value in datesdict.items():
        if value[0] == day_name:
            return value[1]
    return None  # Retorna None se o nome do dia não for encontrado


def show_entries():
    chosen_day = day_dropdown.get()
    stop_games = stop_games_entry.get()
    start_time = start_time_entry.get()
    end_time = end_time_entry.get()
    sort_by_league = sort_by_league_var.get()

    print("Escolha o dia:", chosen_day)
    print("Você quer parar em determinado número de jogos?", stop_games)
    print("Start time:", start_time)
    print("End time:", end_time)
    print("Você quer classificar por liga?", sort_by_league)
    
    default_value = 2  # Valor padrão para quando a string estiver vazia

    try:
        proxies_file = int(stop_games)
    except ValueError:
        proxies_file = default_value

    start_time = datetime.datetime.strptime(start_time, "%H:%M").time()
    end_time = datetime.datetime.strptime(end_time, "%H:%M").time()

    url = {"start_time":start_time, "end_time":end_time}

    leagues = False
    #league_input = input('Do you want to sort by league? (y/n): ')
    if sort_by_league.lower() == 'yes':
        leagues = True
    proxies_file = {'leagues':leagues, 'proxies_file':proxies_file}

    l_de_req=[1, 1000, get_link_by_day(chosen_day,puxa_dias_com_links()), proxies_file, url]
    print("l_de_req ->", l_de_req)

    #show_loading()
    n_threads, run_time, combo_file, proxies_file, url = l_de_req #  get_input()
    print( "n_threads: "+ str(n_threads),"\n", "run_time: "+str(run_time), "\n", "combo_file: "+str(combo_file),"\n","proxies_file: "+ str(proxies_file),"\n", "url: "+ str(url))

    simulation = Simulation(
        url,
        run_time,
        n_threads,
        thread_start_delay,
        proxies_file,
        combo_file
    )

    simulation.start()

# Function to validate the input
def validate_input(P):
    # Only allow numbers (no letters, symbols, etc.)
    if P == "" or P.isdigit():
        return True
    return False


root = ctk.CTk()
root.title("Configurações")
# Set the width and height of the window
width = 820
height = 450

# Define custom colors
bg_color = "grey"
text_color = "white"
light_grey = "#D5D5D5"  # Light grey color
hover_grey = "#C0C0C0"   # Slightly darker grey for hover
#border_color = "#A9A9A9" # Darker grey for 3D effect
dark_color = "#555555"  # Dark grey for the button background
hover_color = "#111111"  # Lighter grey for hover effect
border_color = "#1a1a1a"  # Almost black border to create a sharp contrast
glow_color = "#1f9eff"  # A subtle glowing blue color for hover (optional)

# Set the theme if needed
ctk.set_appearance_mode("dark")  # This might change some colors automatically
ctk.set_default_color_theme("dark-blue")  # Adjust theme as needed

# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate the x and y position to center the window
x_position = int((screen_width - width) / 2)
y_position = int((screen_height - height) / 2)

# Set the geometry with the calculated position
root.geometry(f"{width}x{height}+{x_position}+{y_position}")

bg_image = ctk.CTkImage(dark_image=Image.open("bg4.jpg"), size=(820, 450))

# Create a label to display the image as a background
bg_label = ctk.CTkLabel(root, image=bg_image, text="")  # `text=""` to hide default text
bg_label.place(x=0, y=0, relwidth=1, relheight=1)  # Cover entire window

# Add spacing between widgets
padding_y = 12
padding_round = 20

# Dropdown to choose the day
days = puxa_datas()
padx_value = 60
print(days)

button_font = ctk.CTkFont(family="Helvetica", size=22)
custom_font = ctk.CTkFont(family="Arial", size=14)

# Button with dark color theme, glowing hover, and rounded corners
show_entries_button = ctk.CTkButton(
    root,
    text="Show Entries",
    command=show_entries,
    width=190,
    height=40,
    corner_radius=30,         # Rounded corners for a sleek look
    fg_color=hover_color,      # Dark button background
    hover_color=dark_color,  # Hover color, lighter grey
    border_color=border_color,# Dark border color for contrast
    border_width=3,           # Border width for a sharper effect
    font=button_font,
    cursor="hand2",            # Pointer cursor for better interactivity
    text_color="white"
)

# Set 60px right margin with padx
show_entries_button.pack(side="bottom", anchor="se", padx=(0, 60), pady=(30, 60))

# CTkComboBox for sorting by league
sort_by_league_var = ctk.StringVar(value="YES")
sort_by_league_dropdown = ctk.CTkComboBox(
    root,
    values=["YES", "NO"],
    variable=sort_by_league_var,
    font=custom_font,
    width=190,
    height=30,  # Adjust height for a more rounded shape
    corner_radius=20  # Apply corner radius for rounded appearance
)
sort_by_league_dropdown.pack(side="bottom", anchor="se", padx=(0, padx_value), pady=padding_y)

# Create the CTkEntry widget with rounded corners
end_time_entry = ctk.CTkEntry(
    root,
    placeholder_text="End time (HH:MM)",
    font=custom_font,
    width=190,  # Maintain the same width
    height=30,  # Adjust the height to make it more rounded
    corner_radius=20  # High corner radius to make it rounder (pill shape)
)

# Pack the entry widget into the window
end_time_entry.pack(side="bottom", anchor="se", padx=(0, padx_value), pady=padding_y)

# Create the CTkEntry widget with rounded corners
start_time_entry = ctk.CTkEntry(    
    root,
    placeholder_text="Start time (HH:MM)",
    font=custom_font,
    width=190,  # Maintain the same width
    height=30,  # Adjust the height to make it more rounded
    corner_radius=20  # High corner radius to make it rounder (pill shape)
)

# Pack the entry widget into the window
start_time_entry.pack(side="bottom", anchor="se", padx=(0, padx_value), pady=padding_y)

# CTkEntry for user input (replacing ttk.Entry)
# Create the CTkEntry widget with rounded corners
stop_games_entry = ctk.CTkEntry(
    root,
    placeholder_text="Number of Matches",
    font=custom_font,
    width=190,  # Maintain the same width
    height=30,  # Adjust the height to make it more rounded
    corner_radius=20  # High corner radius to make it rounder (pill shape)
)
stop_games_entry.pack(side="bottom", anchor="se", padx=(0, padx_value), pady=padding_y)

days_var = ctk.StringVar(value=days[0])
day_dropdown = ctk.CTkComboBox(
    root,
    values=days,
    variable=days_var,
    font=custom_font,
    width=190,
    height=30,  # Adjust height for a more rounded shape
    corner_radius=20  # Apply corner radius for rounded appearance
)
day_dropdown.pack(side="bottom", anchor="se", padx=(0, padx_value), pady=(15, padding_y))

root.mainloop()

